#!/usr/bin/env python3
"""
QC: release workbook EVENT INDEX parity.

Codifies the invariant:

    out/Footbag_Results_Release.xlsx sheet "EVENT INDEX" row count
    must equal
    event_results/canonical_input/events.csv row count

The workbook's EVENT INDEX sheet is an identity pass over canonical_input's
events.csv — `build_workbook_release.py::build_event_index` iterates every
row with no filter. Any divergence is a real bug either in the INDEX
builder or in the events dict that feeds it.

## Why NOT compare to out/canonical/events.csv

`pipeline/platform/export_canonical_platform.py` drops disciplines with
coverage_flag=sparse and then drops events with zero remaining disciplines.
out/canonical/events.csv therefore contains more events than canonical_input.
Comparing the workbook against out/canonical would produce a false mismatch.
canonical_input is the workbook's source of truth.

## Exit codes

    0  PASS — counts match
    0  SKIP — release workbook absent, or older than canonical_input/events.csv
           (nothing comparable to compare)
    1  FAIL — counts differ
    2  ERROR — canonical_input/events.csv absent or unreadable
"""

from __future__ import annotations

import csv
import sys
from datetime import datetime, timezone
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
LEGACY_ROOT = SCRIPT_DIR.parents[1]

DEFAULT_CANONICAL_INPUT = LEGACY_ROOT / "event_results" / "canonical_input" / "events.csv"
DEFAULT_WORKBOOK = LEGACY_ROOT / "out" / "Footbag_Results_Release.xlsx"

SHEET_NAME = "EVENT INDEX"
# EVENT INDEX sheet structure:
#   row 1: title row ("EVENT INDEX — ALL EVENTS")
#   row 2: header row
#   rows 3+: data rows (first cell is a four-digit year string)
# Data rows are identified by a first-column value that parses as a year.
_YEAR_MIN = 1970
_YEAR_MAX = 2030


def iso_mtime(path: Path) -> str:
    """The file's modification time as a UTC stamp naming its zone, matching how
    every other operator-facing timestamp in this project is written."""
    stamp = datetime.fromtimestamp(path.stat().st_mtime, tz=timezone.utc)
    return stamp.strftime("%Y-%m-%d %H:%M:%S UTC")


def count_canonical_events(path: Path) -> int:
    with open(path, newline="", encoding="utf-8") as f:
        return sum(1 for _ in csv.DictReader(f))


def count_event_index_rows(workbook_path: Path) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise RuntimeError(f"Workbook missing sheet {SHEET_NAME!r}")
        ws = wb[SHEET_NAME]
        n = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            v = row[0] if row else None
            if v is None:
                continue
            s = str(v).strip()
            if s.isdigit() and _YEAR_MIN <= int(s) <= _YEAR_MAX:
                n += 1
        return n
    finally:
        wb.close()


def main() -> int:
    import argparse
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--canonical-input", type=Path, default=DEFAULT_CANONICAL_INPUT,
                        help=f"canonical_input events.csv (default: {DEFAULT_CANONICAL_INPUT.relative_to(LEGACY_ROOT)})")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK,
                        help=f"release workbook (default: {DEFAULT_WORKBOOK.relative_to(LEGACY_ROOT)})")
    args = parser.parse_args()

    canonical_input = args.canonical_input.resolve()
    workbook = args.workbook.resolve()

    print("=== workbook EVENT INDEX parity ===")
    print(f"canonical_input: {canonical_input}")
    print(f"workbook:        {workbook}")
    print("Invariant: EVENT INDEX row count == canonical_input/events.csv row count.")
    print("(Comparing against out/canonical/events.csv is incorrect — that file")
    print(" intentionally contains more events; sparse-only events are filtered")
    print(" by export_canonical_platform.py before canonical_input.)")

    if not canonical_input.exists():
        print(f"\nERROR: canonical_input not found: {canonical_input}", file=sys.stderr)
        return 2

    try:
        canonical_count = count_canonical_events(canonical_input)
    except Exception as e:
        print(f"\nERROR: failed to read {canonical_input}: {e}", file=sys.stderr)
        return 2

    if not workbook.exists():
        print(f"\nSKIP: release workbook not found (expected at {workbook}).")
        print(f"canonical_input/events.csv rows: {canonical_count}")
        print("No comparison performed.")
        return 0

    # A workbook older than the events file it is an identity pass over is not
    # evidence of anything: the two were built from different canonical sets, so
    # a difference measures their age gap rather than a fault in the builder.
    # This is the normal state of a csv_only build, which rebuilds canonical_input
    # and deliberately leaves the workbook to the mirror pipeline, so failing here
    # would report a defect on every such run for as long as an old workbook sits
    # on disk.
    if workbook.stat().st_mtime < canonical_input.stat().st_mtime:
        print(f"\nSKIP: release workbook is older than {canonical_input.name}.")
        print(f"  workbook:        {workbook} ({iso_mtime(workbook)})")
        print(f"  canonical_input: {canonical_input} ({iso_mtime(canonical_input)})")
        print("The two were built from different canonical sets, so their counts")
        print("are not comparable. Rebuild the workbook through the mirror")
        print("pipeline (./run_pipeline.sh full) to compare them.")
        print(f"canonical_input/events.csv rows: {canonical_count}")
        return 0

    try:
        index_count = count_event_index_rows(workbook)
    except Exception as e:
        print(f"\nERROR: failed to read workbook EVENT INDEX: {e}", file=sys.stderr)
        return 2

    print()
    print(f"canonical_input/events.csv rows: {canonical_count}")
    print(f"workbook EVENT INDEX rows:       {index_count}")

    if canonical_count == index_count:
        print("STATUS: PASS")
        return 0

    delta = index_count - canonical_count
    print(f"STATUS: FAIL — EVENT INDEX differs from canonical_input by {delta:+d}")
    print("The bug is in build_event_index (pipeline/build_workbook_release.py)")
    print("or in what populates its events dict — not in canonical_input.")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
