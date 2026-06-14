#!/usr/bin/env python3
"""Workbook parity QC (run_qc.py optional check ``workbook_qc``).

Hard gate: the release workbook's EVENT INDEX must list exactly the events in
``event_results/canonical_input/events.csv`` (the workbook builder's input),
row-for-row. A mismatch means the builder dropped or invented events relative to
its canonical input.

Informational only (NOT gated): year-sheet event-column count, excluded-sheet
row count, and the partition delta. A strict "EVENT INDEX == year-sheet union +
excluded" partition cannot be computed reliably from the workbook, because year
sheets carry event names (not keys) and some sparse events appear both on a year
sheet and in the excluded list. These counts are printed for visibility; they do
not fail the check.

Invoked by run_qc.py as: ``qc_footbag_workbook.py <workbook.xlsx>``.
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent
DEFAULT_EVENTS_CSV = ROOT / "event_results" / "canonical_input" / "events.csv"
EXCLUDED_SHEET = "QC - EXCLUDED EVENTS"


def _is_event_row(cell_a) -> bool:
    """A data row is one whose first cell starts with a 4-digit year. This skips
    the title, the header, and the trailing legend rows without position math."""
    s = str(cell_a or "").strip()
    return len(s) >= 4 and s[:4].isdigit()


def count_event_index_events(ws) -> int:
    return sum(1 for row in ws.iter_rows(values_only=True) if row and _is_event_row(row[0]))


def count_year_sheet_events(wb) -> int:
    """Year sheets lay events out as columns: row 1 is ``Event`` then one event
    name per column. Sum the non-empty event-name cells across all year sheets."""
    total = 0
    for name in wb.sheetnames:
        if not name.isdigit():
            continue
        first = next(wb[name].iter_rows(min_row=1, max_row=1, values_only=True), ())
        total += sum(1 for c in first[1:] if str(c or "").strip())
    return total


def count_excluded_events(wb) -> int:
    if EXCLUDED_SHEET not in wb.sheetnames:
        return 0
    return sum(1 for row in wb[EXCLUDED_SHEET].iter_rows(values_only=True) if row and _is_event_row(row[0]))


def count_events_csv(path: Path) -> int:
    with Path(path).open(encoding="utf-8", newline="") as fh:
        return sum(1 for _ in csv.DictReader(fh))


def run(workbook_path: Path, events_csv_path: Path = DEFAULT_EVENTS_CSV) -> int:
    workbook_path = Path(workbook_path)
    events_csv_path = Path(events_csv_path)
    if not workbook_path.exists():
        print(f"workbook not found: {workbook_path}", file=sys.stderr)
        return 2
    if not events_csv_path.exists():
        print(
            f"canonical_input events.csv not found: {events_csv_path}. "
            f"Run the platform export (pipeline/platform/export_canonical_platform.py) first.",
            file=sys.stderr,
        )
        return 2

    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    if "EVENT INDEX" not in wb.sheetnames:
        print("workbook has no EVENT INDEX sheet", file=sys.stderr)
        return 2

    event_index = count_event_index_events(wb["EVENT INDEX"])
    events_csv = count_events_csv(events_csv_path)
    year_sheet = count_year_sheet_events(wb)
    excluded = count_excluded_events(wb)

    print("workbook parity QC")
    print(f"  EVENT INDEX events:            {event_index}")
    print(f"  canonical_input/events.csv:    {events_csv}")
    print(f"  [info] year-sheet columns:     {year_sheet}")
    print(f"  [info] excluded-sheet rows:    {excluded}")
    print(f"  [info] partition delta (year+excluded - EVENT INDEX): {year_sheet + excluded - event_index}")

    if event_index != events_csv:
        print(
            f"\nFAIL: EVENT INDEX event count ({event_index}) != "
            f"canonical_input/events.csv row count ({events_csv}). The EVENT INDEX "
            f"must match event_results/canonical_input/events.csv row-for-row. "
            f"Rebuild the workbook (pipeline/build_workbook_release.py) from the "
            f"current canonical_input, or investigate why the builder dropped or "
            f"added events relative to events.csv.",
            file=sys.stderr,
        )
        return 1

    print("\nPASS: EVENT INDEX matches canonical_input/events.csv row-for-row.")
    return 0


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("usage: qc_footbag_workbook.py <workbook.xlsx>", file=sys.stderr)
        return 2
    return run(Path(argv[1]))


if __name__ == "__main__":
    sys.exit(main(sys.argv))
