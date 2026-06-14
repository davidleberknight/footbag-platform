"""Tests for the workbook parity QC (run_qc.py optional check ``workbook_qc``).

Fixture-based: each test builds a minimal release-shaped workbook + a tiny
canonical_input events CSV in tmp_path, so the check is exercised without the
real multi-hour workbook build. The hard gate is EVENT INDEX rows ==
canonical_input/events.csv rows; year-sheet / excluded counts are informational.
"""
import importlib.util
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
MOD = REPO_ROOT / "legacy_data" / "qc_footbag_workbook.py"
_spec = importlib.util.spec_from_file_location("qc_footbag_workbook", MOD)
q = importlib.util.module_from_spec(_spec)
sys.modules["qc_footbag_workbook"] = q
_spec.loader.exec_module(q)


def _make_workbook(path, n_events, year_sheet_events=0, excluded=0):
    wb = openpyxl.Workbook()
    ei = wb.active
    ei.title = "EVENT INDEX"
    ei.append(["EVENT INDEX — ALL EVENTS"])            # title row (not a 4-digit year)
    ei.append(["Year", "Event Name", "Location"])      # header row
    for i in range(n_events):
        ei.append([str(1980 + i % 40), f"Event {i}", "Loc"])
    ei.append(["  NO RESULTS", "legend"])              # trailing legend (not a 4-digit year)
    ys = wb.create_sheet("1995")
    ys.append(["Event"] + [f"E{i}" for i in range(year_sheet_events)])
    ex = wb.create_sheet("QC - EXCLUDED EVENTS")
    ex.append(["QC — EVENTS EXCLUDED"])
    ex.append(["Year", "Event Key", "Event Name"])
    for i in range(excluded):
        ex.append([str(1980 + i), f"k{i}", f"Ex {i}"])
    wb.save(path)


def _make_csv(path, n):
    path.write_text("event_key,year\n" + "".join(f"e{i},2000\n" for i in range(n)), encoding="utf-8")


def test_event_index_count_skips_title_header_legend(tmp_path):
    p = tmp_path / "wb.xlsx"
    _make_workbook(p, n_events=5)
    wb = openpyxl.load_workbook(p, read_only=True)
    assert q.count_event_index_events(wb["EVENT INDEX"]) == 5


def test_year_sheet_and_excluded_counts(tmp_path):
    p = tmp_path / "wb.xlsx"
    _make_workbook(p, n_events=5, year_sheet_events=3, excluded=2)
    wb = openpyxl.load_workbook(p, read_only=True)
    assert q.count_year_sheet_events(wb) == 3
    assert q.count_excluded_events(wb) == 2


def test_events_csv_count(tmp_path):
    p = tmp_path / "events.csv"
    _make_csv(p, 7)
    assert q.count_events_csv(p) == 7


def test_run_passes_when_counts_match(tmp_path, capsys):
    wbp, csvp = tmp_path / "wb.xlsx", tmp_path / "events.csv"
    _make_workbook(wbp, n_events=6)
    _make_csv(csvp, 6)
    assert q.run(wbp, csvp) == 0
    assert "PASS" in capsys.readouterr().out


def test_run_hard_fails_with_actionable_message_on_mismatch(tmp_path, capsys):
    wbp, csvp = tmp_path / "wb.xlsx", tmp_path / "events.csv"
    _make_workbook(wbp, n_events=6)
    _make_csv(csvp, 5)
    assert q.run(wbp, csvp) == 1
    err = capsys.readouterr().err
    assert "FAIL" in err and "6" in err and "5" in err and "row-for-row" in err


def test_run_diagnostics_are_informational_only(tmp_path, capsys):
    # year+excluded != EVENT INDEX, but the run still PASSES on the matching core counts.
    wbp, csvp = tmp_path / "wb.xlsx", tmp_path / "events.csv"
    _make_workbook(wbp, n_events=5, year_sheet_events=3, excluded=99)
    _make_csv(csvp, 5)
    assert q.run(wbp, csvp) == 0
    out = capsys.readouterr().out
    assert "year-sheet columns" in out and "partition delta" in out


def test_run_missing_workbook_returns_2(tmp_path):
    assert q.run(tmp_path / "nope.xlsx", tmp_path / "e.csv") == 2
