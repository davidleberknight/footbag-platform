"""Tests the DATA NOTES - OVERRIDES sheet in build_workbook_release.py.

Fixture-based: a fresh workbook + tiny override files + a small events/legacy-id
map. Pins that results-file swaps and field overrides are listed (with source +
note), exclusions are skipped (they live on QC - EXCLUDED EVENTS), and unmapped
event_ids fall back to a visible placeholder rather than vanishing.
"""
import importlib.util
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
MOD = REPO_ROOT / "legacy_data" / "pipeline" / "build_workbook_release.py"
_spec = importlib.util.spec_from_file_location("build_workbook_release", MOD)
b = importlib.util.module_from_spec(_spec)
sys.modules["build_workbook_release"] = b
_spec.loader.exec_module(b)


def test_override_notes_lists_overrides_skips_exclusions_and_handles_unmapped(tmp_path):
    rcsv = tmp_path / "r.csv"
    rcsv.write_text("event_id,file,replace,notes\n111,f.txt,true,typo fix\n", encoding="utf-8")
    ejsonl = tmp_path / "e.jsonl"
    ejsonl.write_text(
        '{"event_id":"222","source":"stage2.LOCATION_OVERRIDES","location":"X","confidence":"high"}\n'
        '{"event_id":"333","source":"manual","exclude":true}\n'
        '{"event_id":"444","source":"qc_review","event_type":"social","reason":"club social"}\n',
        encoding="utf-8",
    )
    events = {
        "k1": {"year": "1999", "event_name": "Western Regional", "legacy_event_id": "111"},
        "k2": {"year": "2001", "event_name": "Frankfurt Open", "legacy_event_id": "222"},
        # event 444 intentionally absent → unmapped fallback
    }
    legacy_map = {"111": "k1", "222": "k2"}

    wb = openpyxl.Workbook()
    b.build_override_notes(wb, events, legacy_map, rcsv, ejsonl)
    ws = wb["DATA NOTES - OVERRIDES"]
    rows = list(ws.iter_rows(values_only=True))
    data = [r for r in rows if r[2] in ("results-file", "field")]

    assert sorted(r[2] for r in data) == ["field", "field", "results-file"]   # exclusion 333 skipped

    rf = next(r for r in data if r[2] == "results-file")
    assert rf[0] == "1999" and rf[1] == "Western Regional" and rf[4] == "typo fix"

    loc = next(r for r in data if r[3] == "stage2.LOCATION_OVERRIDES")
    assert loc[1] == "Frankfurt Open" and "location" in loc[4]

    unmapped = next(r for r in data if r[3] == "qc_review")
    assert "not in index" in unmapped[1] and "club social" in unmapped[4]

    assert all("333" not in str(c) for r in rows for c in r)   # exclusion absent everywhere


def test_override_notes_empty_when_no_files(tmp_path):
    wb = openpyxl.Workbook()
    b.build_override_notes(wb, {}, {}, tmp_path / "none.csv", tmp_path / "none.jsonl")
    ws = wb["DATA NOTES - OVERRIDES"]
    assert [r for r in ws.iter_rows(values_only=True) if r[2] in ("results-file", "field")] == []
