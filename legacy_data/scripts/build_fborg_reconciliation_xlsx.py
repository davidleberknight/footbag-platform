"""
Build an XLSX reconciliation workbook between Footbag.org and the IFPA
freestyle dictionary, with evidence-density scoring and source-lineage
tracking.

Read-only. Composes from existing artifacts:
  - legacy_data/reports/freestyle_dict_coverage_diff.csv  (corpus rows)
  - legacy_data/reports/structural_alias_adjudication.csv (policy verdict)
  - database/footbag.db                                   (canonical + media)

Output:
  - legacy_data/reports/fborg_reconciliation.xlsx
    - sheet "Reconciliation": one row per corpus entry + evidence/lineage
    - sheet "Summary":         routing counts + color legend
    - sheet "Evidence Summary": evidence-shape distributions

Architecture (source-agnostic):

  The SOURCES registry below lists every external/internal evidence
  source the workbook tracks. Adding a future source (e.g. when the
  ~569-trick Footbagmoves.com corpus arrives) is an entry append +
  extending the presence-detection in `evaluate_evidence` — no
  schema/column-shape redesign needed. Each source contributes a
  presence flag column (`evidence_<id>`) and feeds into lineage and
  scoring.

Routing (evidence-aware, 2026-05-09 revision):

  A row is routed based on (confidence band, conflict signal):
    HIGH/MEDIUM + no conflict   → auto_resolved | james_reviewable
    HIGH/MEDIUM + conflict      → needs_red
    LOW         + no conflict   → provisional_candidate
    LOW         + conflict      → quarantine

  `provisional_candidate` and `quarantine` are new states that prevent
  bare-evidence corpus rows from inflating the Red queue. James-
  pre-decided overrides bypass routing entirely.

  Footbag.org rows are valid external evidence by default. Missing-
  from-IFPA never alone implies low validity — it implies coverage
  gap, which the evidence band makes explicit.

Does not modify any ontology data, CSVs, or DB.
"""
import argparse
import csv
import re
import sqlite3
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
DEFAULT_DIFF        = REPO / "legacy_data" / "reports" / "freestyle_dict_coverage_diff.csv"
DEFAULT_ADJUDICATION = REPO / "legacy_data" / "reports" / "structural_alias_adjudication.csv"
DEFAULT_DB          = REPO / "database" / "footbag.db"
DEFAULT_OUT         = REPO / "legacy_data" / "reports" / "fborg_reconciliation.xlsx"

# James-pre-decided overrides. Match on lowercased external_name.
# Each entry: lowercased_name → (decision, comment).
JAMES_PRE_DECIDED: dict[str, tuple[str, str]] = {
    "triple around the world":     ("valid productive variant (not standalone canonical)",
                                    "James direct adjudication: treated as a valid productive variant of around-the-world, not a standalone canonical named trick under current dictionary policy (§10)."),
    "double around the world heel":("valid productive variant (not standalone canonical)",
                                    "James direct adjudication: treated as a valid productive variant of around-the-world (with heel-stall ending), not a standalone canonical named trick under current dictionary policy (§10)."),
    "blender":                     ("canonical (existing)",
                                    "James direct adjudication: blender is an active canonical row in IFPA dictionary."),
    "butterfly swirl":             ("canonical-candidate (provisional)",
                                    "James direct adjudication: community-named compound; promote to canonical pending notation/ADD review."),
    "barfly":                      ("canonical (existing)",
                                    "James direct adjudication: barfly is an active canonical row in IFPA dictionary."),
    "down double down":            ("blocked by down-family policy",
                                    "James direct adjudication: defer until down-family policy ruling — see red-followup-down-family-2026-05.md."),
}


# ─── Evidence model ─────────────────────────────────────────────────────────
#
# `SOURCES` is the source-agnostic registry. Each entry: (id, label,
# strength). `strength` is the lineage priority used to pick
# `strongest_source` — higher wins. It is INDEPENDENT of scoring weight.
# To onboard a future source (Footbagmoves.com, Erik Chan's Foundations,
# Polini's Pointers, etc.), append one entry here and extend
# `evaluate_evidence` to populate its presence flag.
SOURCES: list[tuple[str, str, int]] = [
    ("fborg",     "Footbag.org",                    10),
    ("anztrikz",  "Anz' Trikz",                     20),
    ("passback",  "Passback records",               30),
    ("tt",        "Tricks of the Trade",            40),
    ("ifpa",      "IFPA dictionary (canonical)",    50),
    ("red",       "Red endorsement",                60),
    # Future sources go here, e.g.:
    # ("fbmoves",  "Footbagmoves.com",               25),
]
SOURCE_IDS = [s[0] for s in SOURCES]
SOURCE_LABEL = {s[0]: s[1] for s in SOURCES}
SOURCE_STRENGTH = {s[0]: s[2] for s in SOURCES}

# Scoring weights — independent of source registry. The keys here are
# evidence kinds (some align 1:1 with source ids; others — like
# `structural_parse`, `jobs_notation`, `video`, `multiple_sources` —
# are derived signals).
SCORE_WEIGHTS: dict[str, int] = {
    "ifpa":              5,   # existing IFPA canonical match
    "structural_parse":  3,   # structural decomposition resolves
    "jobs_notation":     2,   # row carries Jobs notation
    "tt":                3,   # TT video evidence
    "passback":          2,   # passback record evidence
    "red":               5,   # Red-endorsed (review_status='expert_reviewed')
    "video":             1,   # any video media linked
    "multiple_sources":  2,   # presence count ≥ 2
}

# Confidence band thresholds (against `evidence_total_score`). Tuned so
# that matched-canonical + Jobs notation crosses into HIGH and bare
# name-only rows fall into LOW. Mid-band catches documented but
# unverified rows (notation present, no corroboration).
CONFIDENCE_HIGH_MIN   = 7
CONFIDENCE_MEDIUM_MIN = 2

CONFIDENCE_LABELS = ("LOW", "MEDIUM", "HIGH")

# ─── R1 sheet schemas ────────────────────────────────────────────────────────
#
# Phase R1 (2026-05-10) introduces three additional sheets alongside the
# existing Reconciliation / Summary / Evidence Summary. Schemas are pinned
# here as module-level constants so they're easy to review in one place.
#
# Authority layering across the Tricks sheet's column groups:
#   Identity / Editorial truth  → cols 1–7
#   Notation                    → cols 8–12   (semantic vs operational vs conflict)
#   Parser-derived              → cols 13–15
#   Media                       → cols 16–18
#   External evidence           → cols 19–22  (fborg, fmoves)
#   Conflict + decision + review → cols 23–28
#   Audit                       → cols 29–30
#
# R1 populates: Identity, Editorial, Notation (semantic only), Parser, Media,
# External (fborg only), Adoption, Red review status. Reserved-for-later
# columns render with sentinel defaults (operational notation, fmoves match,
# conflict columns). Schema reservation now means R2/R4 don't need to
# re-shape the workbook.

# Source-tier registry for the Tricks sheet's tutorial_status / media_count
# computation. Mirrors src/services/freestyleService.ts SOURCE_TIER (post-
# Phase 2b: shred_global is DEMONSTRATION). Kept small + explicit so it's
# easy to keep in sync with the platform service when sources move tiers.
TUTORIAL_TIER_SOURCES: frozenset[str] = frozenset({
    "tt_youtube", "footbagspot_tutorials", "polini_pointers",
    "footbag_foundations", "everything_footbag",
    "anz_trikz", "footbagspot_passback",  # held at TUTORIAL pending Phase 2d review
})
DEMO_TIER_SOURCES: frozenset[str] = frozenset({
    "shred_global", "footbag_finland", "flipsider_footbag",
})

# Tricks sheet — one row per active freestyle_tricks row.
# (column_name, width). Visual grouping is by intent; see authority-layer
# comment above for which span belongs to which group.
TRICKS_COLUMNS: list[tuple[str, int]] = [
    # Identity (1–6)
    ("slug",                          22),
    ("display_name",                  24),
    ("add",                            5),
    ("family",                        18),
    ("category",                      11),
    ("aliases_count",                  7),
    # Editorial (7)
    ("description",                   48),
    # Notation (8–12)
    ("semantic_notation",             32),
    ("semantic_notation_status",      18),
    ("operational_notation",          28),
    ("operational_notation_status",   20),
    ("notation_conflict",             20),
    # Parser-derived (13–15)
    ("parser_status",                 22),
    ("parser_computed_add",            7),
    ("add_conflict",                  18),
    # Media (16–18)
    ("tutorial_status",               18),
    ("media_count",                    7),
    ("record_count",                   7),
    # External evidence (19–22)
    ("fborg_match",                   18),
    ("fborg_name",                    24),
    ("fborg_add",                      6),
    ("fmoves_match",                  18),
    # Conflict + decision + review (23–28)
    ("ontology_conflict",             20),
    ("adoption_status",               14),
    ("james_decision",                30),
    ("red_review_needed",              8),
    ("red_review_topic",              18),
    ("red_review_id",                  6),
    # Audit (29–30)
    ("page_completeness",             14),
    ("notes",                         40),
]

# Aliases sheet — one row per alias, joined to canonical via trick_slug.
ALIASES_COLUMNS: list[tuple[str, int]] = [
    ("canonical_slug",                22),
    ("alias",                         28),
    ("alias_source",                  16),
    ("display_priority",               7),
    ("confidence",                    11),
    ("notes",                         40),
]

# Red Review Queue sheet — pending adjudication items. R1 carries a small
# representative seed (~6–8 rows) hand-picked from OPEN_QUESTIONS.md and
# red-followup-*.md. R3 replaces the seed with a file-backed reader.
RED_QUEUE_COLUMNS: list[tuple[str, int]] = [
    ("id",                             5),
    ("topic",                         18),
    ("slug_affected",                 24),
    ("claim",                         60),
    ("evidence_links",                40),
    ("proposed_resolution",           60),
    ("red_status",                    14),
    ("red_response",                  40),
    ("decision_date",                 12),
    ("notes",                         30),
]

# Page-completeness band thresholds (Tricks column 29). Pure-derived from
# existing fields; populated in R1. Reader scans this single column to see
# where each canonical sits on a 5-band readiness ladder.
PAGE_COMPLETENESS_BANDS = ("minimal", "educational", "media-backed", "advanced", "showcase")


def parse_int(value: str | None) -> int | None:
    if not value: return None
    try: return int(value)
    except (ValueError, TypeError): return None


def slugify(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return s or "unknown"


def load_canonical(db: sqlite3.Connection) -> dict[str, dict]:
    """slug → {name, adds, notation, description, is_active, review_status}."""
    rows = db.execute(
        """SELECT slug, canonical_name, adds, notation, description, is_active, review_status
           FROM freestyle_tricks"""
    ).fetchall()
    out = {}
    for slug, name, adds, notation, descr, is_active, status in rows:
        out[slug] = {
            "name":          name,
            "adds":          parse_int(adds) if adds else None,
            "notation":      notation or "",
            "description":   descr or "",
            "is_active":     bool(is_active),
            "review_status": status or "",
        }
    return out


# Map from media-system source_id (as stored in freestyle_media_assets)
# to evidence-source registry id. Adding a new media source means adding
# its source_id here. Sources not in this map don't contribute to
# per-source presence flags but DO still light `evidence_video`.
MEDIA_SOURCE_MAP: dict[str, str] = {
    "tt_youtube":            "tt",
    "anz_trikz":             "anztrikz",
    "passback_records":      "passback",
    # Future expansion goes here, e.g.:
    # "fbmoves":               "fbmoves",
}


def load_trick_source_index(db: sqlite3.Connection) -> dict[str, set[str]]:
    """slug → set of evidence-source ids that have a media link to it.
    Source ids are mapped through MEDIA_SOURCE_MAP; unmapped media
    source_ids contribute only via `evidence_video`."""
    out: dict[str, set[str]] = {}
    rows = db.execute("""
        SELECT DISTINCT l.entity_id, a.source_id
        FROM freestyle_media_links l
        JOIN freestyle_media_assets a ON a.id = l.media_id
        WHERE l.entity_type = 'trick'
    """).fetchall()
    for slug, source_id in rows:
        out.setdefault(slug, set())
        ev_id = MEDIA_SOURCE_MAP.get(source_id)
        if ev_id:
            out[slug].add(ev_id)
    return out


def load_trick_video_set(db: sqlite3.Connection) -> set[str]:
    """slugs that have at least one video media link (any source).
    Independent of MEDIA_SOURCE_MAP — drives evidence_video."""
    rows = db.execute("""
        SELECT DISTINCT l.entity_id
        FROM freestyle_media_links l
        JOIN freestyle_media_assets a ON a.id = l.media_id
        WHERE l.entity_type = 'trick' AND a.media_type = 'video'
    """).fetchall()
    return {r[0] for r in rows}


# ─── R1 loaders ──────────────────────────────────────────────────────────────

def _classify_page_completeness(
    description: str, family: str, semantic_notation: str,
    parser_status: str, tutorial_status: str, media_count: int,
    record_count: int, operational_notation: str, review_status: str,
) -> str:
    """Pure-derived 5-band readiness classification — see PAGE_COMPLETENESS_BANDS.
    R1 surfaces only minimal/educational/media-backed/advanced; showcase
    requires operational_notation populated (R4)."""
    has_description = bool(description and description.strip())
    has_family      = bool(family)
    has_notation    = bool(semantic_notation and semantic_notation.strip())
    has_parser      = parser_status in {"exact_modifier_derived", "exact_self_atom"}
    has_media       = media_count > 0 or record_count > 0 or tutorial_status != "no-video"
    has_op_notation = bool(operational_notation and operational_notation.strip())
    is_red_endorsed = review_status == "expert_reviewed"

    if has_op_notation and has_notation and has_parser and has_media and is_red_endorsed and media_count >= 2:
        return "showcase"
    if has_notation and has_parser and has_media:
        return "advanced"
    if has_description and has_family and has_media:
        return "media-backed"
    if has_description and has_family:
        return "educational"
    return "minimal"


def load_trick_index_for_workbook(db: sqlite3.Connection) -> list[dict]:
    """One row per active freestyle_tricks row. Joined to per-trick aggregates
    needed by the Tricks sheet: alias count, media counts by tier, record
    count, parser status. Reads from BOTH legacy (freestyle_media_*) and
    curator (media_items + media_tags) channels for media-coverage counts.
    Read-only."""

    base_rows = db.execute("""
        SELECT slug, canonical_name, adds, base_trick, trick_family, category,
               description, notation, is_active, review_status,
               jobs_notation_raw, computed_adds, add_formula_status
        FROM freestyle_tricks
        WHERE is_active = 1
        ORDER BY CAST(adds AS INTEGER), slug
    """).fetchall()

    alias_counts: dict[str, int] = dict(db.execute("""
        SELECT trick_slug, COUNT(*) FROM freestyle_trick_aliases GROUP BY trick_slug
    """).fetchall())

    media_counts: dict[str, int] = {}
    tutorial_slugs: set[str] = set()
    demo_slugs:     set[str] = set()

    # Channel 1: legacy freestyle_media_*
    for slug, source_id in db.execute("""
        SELECT l.entity_id, a.source_id
        FROM freestyle_media_links l
        JOIN freestyle_media_assets a ON a.id = l.media_id
        WHERE l.entity_type = 'trick' AND a.is_active = 1
    """).fetchall():
        media_counts[slug] = media_counts.get(slug, 0) + 1
        if source_id in TUTORIAL_TIER_SOURCES:
            tutorial_slugs.add(slug)
        elif source_id in DEMO_TIER_SOURCES:
            demo_slugs.add(slug)

    # Channel 2: curator-tagged media_items + media_tags + tags + freestyle_tricks
    for slug, source_id in db.execute("""
        SELECT ft.slug, mi.source_id
        FROM media_items mi
        JOIN media_tags mt ON mt.media_id = mi.id
        JOIN tags t ON t.id = mt.tag_id
        JOIN freestyle_tricks ft ON ('#' || ft.slug) = t.tag_normalized
        WHERE mi.moderation_status = 'active'
          AND mi.source_id IS NOT NULL
          AND ft.is_active = 1
    """).fetchall():
        media_counts[slug] = media_counts.get(slug, 0) + 1
        if source_id in TUTORIAL_TIER_SOURCES:
            tutorial_slugs.add(slug)
        elif source_id in DEMO_TIER_SOURCES:
            demo_slugs.add(slug)

    # Records (probable + verified only — public-filter parity)
    record_counts: dict[str, int] = dict(db.execute("""
        SELECT LOWER(REPLACE(trick_name,' ','-')) AS slug, COUNT(*)
        FROM freestyle_records
        WHERE confidence IN ('probable','verified')
          AND superseded_by IS NULL
        GROUP BY slug
    """).fetchall())

    out: list[dict] = []
    for row in base_rows:
        slug, canonical_name, adds, base_trick, trick_family, category, \
            description, notation, is_active, review_status, \
            jobs_notation_raw, computed_adds, add_formula_status = row

        # Adoption status (derived from is_active + review_status)
        if is_active and review_status == "expert_reviewed":
            adoption = "live"
        elif is_active and review_status in ("pending", "curated"):
            adoption = "draft"
        elif not is_active:
            adoption = "hidden"
        else:
            adoption = "live"

        # Tutorial status (post-Phase 2 taxonomy)
        if slug in tutorial_slugs:
            tutorial_status = "tutorial-coverage"
        elif slug in demo_slugs:
            tutorial_status = "demo-only"
        else:
            tutorial_status = "no-video"

        # Semantic-notation authoring status
        if notation and notation.strip():
            notation_status = "authored"
        elif add_formula_status in ("exact_modifier_derived", "exact_self_atom"):
            notation_status = "parser-derived"
        else:
            notation_status = "pending"

        page_completeness = _classify_page_completeness(
            description=description or "", family=trick_family or "",
            semantic_notation=notation or "",
            parser_status=add_formula_status or "",
            tutorial_status=tutorial_status,
            media_count=media_counts.get(slug, 0),
            record_count=record_counts.get(slug, 0),
            operational_notation="",  # R1: always empty
            review_status=review_status or "",
        )

        out.append({
            # Identity
            "slug":                       slug,
            "display_name":               canonical_name,
            "add":                        parse_int(adds) if adds else "",
            "family":                     trick_family or "",
            "category":                   category or "",
            "aliases_count":              alias_counts.get(slug, 0),
            # Editorial
            "description":                description or "",
            # Notation
            "semantic_notation":          notation or "",
            "semantic_notation_status":   notation_status,
            "operational_notation":       "",                # R4
            "operational_notation_status":"not-applicable",  # R4
            "notation_conflict":          "none",            # R2
            # Parser-derived
            "parser_status":              add_formula_status or "not-run",
            "parser_computed_add":        computed_adds if computed_adds is not None else "",
            "add_conflict":               "none",            # R2
            # Media
            "tutorial_status":            tutorial_status,
            "media_count":                media_counts.get(slug, 0),
            "record_count":               record_counts.get(slug, 0),
            # External evidence (R1: fborg only; populated downstream from diff_rows)
            "fborg_match":                "",
            "fborg_name":                 "",
            "fborg_add":                  "",
            "fmoves_match":               "no_external_match",  # R4
            # Conflict + decision + review
            "ontology_conflict":          "none",            # R2
            "adoption_status":            adoption,
            "james_decision":             "",                # R5
            "red_review_needed":          "FALSE",           # populated from queue join
            "red_review_topic":           "",
            "red_review_id":              "",
            # Audit
            "page_completeness":          page_completeness,
            "notes":                      "",
        })

    return out


def load_aliases_for_workbook(db: sqlite3.Connection) -> list[dict]:
    """One row per alias on an active trick. Sorted by canonical_slug then
    alias for predictable scan order."""
    rows = db.execute("""
        SELECT a.trick_slug, a.alias_text, a.alias_type
        FROM freestyle_trick_aliases a
        JOIN freestyle_tricks t ON t.slug = a.trick_slug
        WHERE t.is_active = 1
        ORDER BY a.trick_slug, a.alias_text
    """).fetchall()
    out: list[dict] = []
    for trick_slug, alias_text, alias_type in rows:
        out.append({
            "canonical_slug":   trick_slug,
            "alias":            alias_text or "",
            "alias_source":     "curated",
            "display_priority": "",
            "confidence":       "HIGH",
            "notes":            (alias_type or ""),
        })
    return out


def seed_red_queue() -> list[dict]:
    """Hand-picked representative Red review queue (R1 seed). Lifted from
    OPEN_QUESTIONS.md §2 + red-followup-down-family-2026-05.md.
    R3 replaces this with a file-backed reader."""
    return [
        {
            "id":                  1,
            "topic":               "add_dispute",
            "slug_affected":       "royale",
            "claim":               "royale ADD value uncertain. Red pt4 replied '?' (will look into); row remains pending+inactive.",
            "evidence_links":      "OPEN_QUESTIONS.md §2 HIGH",
            "proposed_resolution": "Hold pending; re-ask in next Red packet.",
            "red_status":          "pending",
            "red_response":        "",
            "decision_date":       "",
            "notes":               "",
        },
        {
            "id":                  2,
            "topic":               "decomposition",
            "slug_affected":       "eggbeater",
            "claim":               "Construction conflict: Red pt2 says Eggbeater = Atomic Legover; current row description says 'Illusion-modified legover'. Either description is wrong or atomic/illusioning are interchangeable in legover branch.",
            "evidence_links":      "OPEN_QUESTIONS.md §2 HIGH; red-corrections-pt2.txt",
            "proposed_resolution": "Confirm atomic-legover form with Red; correct description.",
            "red_status":          "pending",
            "red_response":        "",
            "decision_date":       "",
            "notes":               "",
        },
        {
            "id":                  3,
            "topic":               "new_trick",
            "slug_affected":       "flail",
            "claim":               "flail and omelette canonical mappings — Red pt1 said both 'should be added'. footbag.org suggests Symposium Illusion (flail) and Atomic Illusion (omelette). Red pt4 silent.",
            "evidence_links":      "OPEN_QUESTIONS.md §2 HIGH; red-corrections-pt1.txt",
            "proposed_resolution": "Confirm canonical forms + ADD values; promote to active.",
            "red_status":          "pending",
            "red_response":        "",
            "decision_date":       "",
            "notes":               "Pair adjudication; same source ontology branch.",
        },
        {
            "id":                  4,
            "topic":               "add_dispute",
            "slug_affected":       "blistering",
            "claim":               "Blistering existence + ADD value. Red pt9 silently skipped this when answering Sailing in same batch. Sailing resolved (Pixie Quantum equivalence, 2 ADD); Blistering remains unresolved.",
            "evidence_links":      "OPEN_QUESTIONS.md §1B + §2 HIGH; red_corrections_pt9.csv",
            "proposed_resolution": "Re-ask in next Red packet.",
            "red_status":          "pending",
            "red_response":        "",
            "decision_date":       "",
            "notes":               "",
        },
        {
            "id":                  5,
            "topic":               "family",
            "slug_affected":       "down",
            "claim":               "Down-family deferred at canonical layer (down / double-down / plant patterns). External corpus has rows; cannot resolve until policy unblocks.",
            "evidence_links":      "red-followup-down-family-2026-05.md",
            "proposed_resolution": "Policy decision needed before any down-* row promotes.",
            "red_status":          "pending",
            "red_response":        "",
            "decision_date":       "",
            "notes":               "Policy-blocked; affects multiple external rows.",
        },
        {
            "id":                  6,
            "topic":               "category",
            "slug_affected":       "pogo",
            "claim":               "Pogo set handling. Red pt2: Pogo is a set, no ADD bonus. Many pogo-* rows are pending; decide whether to mark non-scoring set entries or leave pending.",
            "evidence_links":      "OPEN_QUESTIONS.md §2 LOW",
            "proposed_resolution": "Promote pogo-* rows with ADD=0 marker, OR keep pending until usage emerges.",
            "red_status":          "pending",
            "red_response":        "",
            "decision_date":       "",
            "notes":               "Low priority; nice-to-have.",
        },
        {
            "id":                  7,
            "topic":               "notation_form",
            "slug_affected":       "sumo",
            "claim":               "Notation form for Sumo (= Nuclear Mirage). Pedagogical-clarity name-form 'SUMO' currently authored; alternative 'NUCLEAR MIRAGE' would expose the structural decomposition.",
            "evidence_links":      "NOTATION_STYLE_GUIDE §5.7; red_corrections_pt9.csv",
            "proposed_resolution": "Keep SUMO (humans-first per §0); decomposition stays in description + editorial layer.",
            "red_status":          "pending",
            "red_response":        "",
            "decision_date":       "",
            "notes":               "Already-applied per §5.7 default; flagging for visibility.",
        },
    ]


def evaluate_evidence(
    diff_row: dict,
    adj_row: dict | None,
    canonical_row: dict | None,
    trick_sources: dict[str, set[str]],
    video_set: set[str],
) -> dict:
    """Compute the full evidence + lineage record for one corpus row.

    Returns a dict with:
      evidence_<id> ('yes'/'no') for each id in SOURCE_IDS
      evidence_jobs_notation, evidence_structural_parse,
      evidence_existing_ifpa, evidence_video, evidence_multiple_sources
      evidence_total_score (int)
      evidence_confidence ('LOW'/'MEDIUM'/'HIGH')
      first_seen_source (corpus driver — fborg today; future: fbmoves)
      strongest_source  (highest-strength source where evidence present)
      supporting_sources (comma-joined list, strength-desc order)
    """
    canonical_match = (diff_row.get("canonical_match") == "yes")
    canonical_slug  = diff_row.get("canonical_slug") or ""
    external_slug   = diff_row.get("external_slug") or ""
    target_slugs    = {s for s in (canonical_slug, external_slug) if s}

    # Collect all evidence-source ids present on any candidate slug.
    sources_present: set[str] = set()
    for slug in target_slugs:
        sources_present |= trick_sources.get(slug, set())

    # fborg presence — every corpus row in this driver is a Footbag.org row.
    sources_present.add("fborg")
    # ifpa presence — matched against an active canonical row.
    if canonical_match:
        sources_present.add("ifpa")
    # Red endorsement — via canonical row's review_status.
    if canonical_match and canonical_row and canonical_row.get("review_status") == "expert_reviewed":
        sources_present.add("red")
    # Passback evidence — also light if the diff CSV flags `has_records=yes`,
    # which catches record evidence even when no media row exists.
    if diff_row.get("has_records") == "yes":
        sources_present.add("passback")

    # Derived signals.
    has_jobs_notation     = diff_row.get("has_external_notation") == "yes"
    has_video             = (
        diff_row.get("has_media") == "yes" or
        diff_row.get("media_present") == "yes" or
        any(slug in video_set for slug in target_slugs)
    )
    auto_classification   = (adj_row or {}).get("auto_classification", "") or ""
    structural_parse_ok   = auto_classification in (
        "STRUCTURAL-ALIAS-FULL",
        "STRUCTURAL-ALIAS-VIA-ALTNAME",
    )

    # Multiple-sources flag (≥ 2 evidence-source ids present).
    multiple_sources = len(sources_present) >= 2

    # Score: presence flags weighted per SCORE_WEIGHTS.
    score = 0
    if "ifpa"     in sources_present: score += SCORE_WEIGHTS["ifpa"]
    if structural_parse_ok:           score += SCORE_WEIGHTS["structural_parse"]
    if has_jobs_notation:             score += SCORE_WEIGHTS["jobs_notation"]
    if "tt"       in sources_present: score += SCORE_WEIGHTS["tt"]
    if "passback" in sources_present: score += SCORE_WEIGHTS["passback"]
    if "red"      in sources_present: score += SCORE_WEIGHTS["red"]
    if has_video:                     score += SCORE_WEIGHTS["video"]
    if multiple_sources:              score += SCORE_WEIGHTS["multiple_sources"]

    # Confidence band.
    if score >= CONFIDENCE_HIGH_MIN:
        confidence = "HIGH"
    elif score >= CONFIDENCE_MEDIUM_MIN:
        confidence = "MEDIUM"
    else:
        confidence = "LOW"

    # Lineage.
    first_seen_source = "fborg"   # corpus driver; will branch when fbmoves ingestion lands
    present_sorted = sorted(sources_present, key=lambda s: -SOURCE_STRENGTH.get(s, 0))
    strongest_source = present_sorted[0] if present_sorted else ""
    supporting_sources = ", ".join(present_sorted)

    record = {
        "evidence_total_score":        score,
        "evidence_confidence":         confidence,
        "evidence_jobs_notation":      "yes" if has_jobs_notation else "no",
        "evidence_structural_parse":   "yes" if structural_parse_ok else "no",
        "evidence_existing_ifpa":      "yes" if "ifpa" in sources_present else "no",
        "evidence_video":              "yes" if has_video else "no",
        "evidence_multiple_sources":   "yes" if multiple_sources else "no",
        "first_seen_source":           first_seen_source,
        "strongest_source":            strongest_source,
        "supporting_sources":          supporting_sources,
    }
    for sid in SOURCE_IDS:
        record[f"evidence_{sid}"] = "yes" if sid in sources_present else "no"
    return record


def load_adjudication(path: Path) -> dict[str, dict]:
    """slug → adjudication row dict."""
    out = {}
    with path.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            out[r["slug"]] = r
    return out


def derive_recon_status(diff_row: dict, adj_row: dict | None) -> tuple[str, str]:
    """Return (reconciliation_status, comment). Maps adjudicator + diff
    state into a single status field plus a human comment.

    Policy revision (2026-05-09): Footbag.org rows are valid external
    evidence by default. Unmatched rows get `fborg_supported_missing`
    rather than the older `external_only` / `fborg_only` framing,
    UNLESS a specific structural verdict (productive-multiplicity,
    structural-alias, policy-block, ADD-conflict, etc.) applies.
    """
    match_type = diff_row.get("match_type", "")
    if match_type == "exact canonical":
        return ("matched_canonical",
                "Footbag.org name matches an active IFPA canonical row.")
    if match_type == "likely structural alias":
        slug = diff_row.get("canonical_slug") or "?"
        return ("structural_alias_inline",
                f"Footbag.org name decomposes structurally to canonical `{slug}`; render as Common alias on the canonical row in Deep Dive.")
    if match_type != "no match":
        return ("unknown_match_type", f"diff match_type={match_type!r}")

    # match_type = "no match" → use adjudicator verdict, falling through
    # to fborg_supported_missing when no specific structural verdict applies.
    cls = (adj_row or {}).get("auto_classification", "")
    decomposition = (adj_row or {}).get("decomposition", "") or ""
    reason = (adj_row or {}).get("reason", "") or ""

    if cls == "PRODUCTIVE-MULTIPLICITY-NON-CANONICAL":
        return ("productive_multiplicity_non_canonical",
                f"Productive multiplicity ({decomposition}); per CANONICALIZATION_POLICY §10 default, no canonical row earned. Treated as a valid productive variant.")
    if cls == "STRUCTURAL-ALIAS-FULL":
        return ("structural_alias_full",
                f"Decomposition: {decomposition}. Render as alias on the decomposition's canonical.")
    if cls == "STRUCTURAL-ALIAS-VIA-ALTNAME":
        return ("structural_alias_via_altname",
                f"External name is community shorthand; alt-name documents structural recipe ({decomposition}).")
    if cls == "STRUCTURAL-ALIAS-ADD-CONFLICT":
        return ("add_conflict",
                f"Decomposition resolves but ADD math disagrees: {decomposition}. {reason}")
    if cls == "POLICY-BLOCKED-DOWN":
        return ("policy_blocked_down",
                f"Down-family policy not yet settled. Deferred. {reason}")
    if cls == "PRODUCTIVE-MULTIPLICITY-BASE-UNRESOLVABLE":
        return ("productive_multiplicity_base_unresolved",
                f"Count-prefix pattern but base unresolved: {decomposition}.")
    if cls == "CANONICAL-CANDIDATE-NAMED-IDENTITY":
        return ("canonical_candidate",
                f"Strong evidence ({(adj_row or {}).get('evidence_score','?')}); community-named identity; pending dictionary-policy decision on canonical elevation.")

    # Default for unmatched rows (including EXTERNAL-ONLY-DESCRIPTIVE,
    # UNRESOLVED, and rows with no adjudication entry): valid external
    # evidence missing from IFPA. James-reviewable.
    return ("fborg_supported_missing",
            "Documented on Footbag.org but not yet present in the IFPA dictionary. Treated as valid external evidence; James-reviewable for whether to add a canonical / alias / leave-as-external entry.")


def _normalize_name_key(name: str) -> str:
    """Lowercase + drop hyphens / whitespace runs so 'Down Double-Down'
    and 'down double down' collapse to the same key."""
    return re.sub(r"[\s\-]+", " ", (name or "").lower()).strip()


# Reconciliation statuses that constitute a confirmed conflict signal —
# these route to Red. Everything else under match_type=no match is
# treated as valid external evidence and routes to james_reviewable.
CONFIRMED_CONFLICT_STATUSES = {
    "add_conflict",                              # ADD math disagrees
    "policy_blocked_down",                       # dictionary policy uncertainty
    "productive_multiplicity_base_unresolved",   # base-canonical question
    "canonical_candidate",                       # canonical-elevation question
}


AUTO_RESOLVED_STATUSES = {
    "matched_canonical",
    "structural_alias_inline",
    "productive_multiplicity_non_canonical",
    "structural_alias_full",
    "structural_alias_via_altname",
}


def adjudication_for(
    name: str,
    recon_status: str,
    diff_row: dict,
    adj_row: dict | None,
    evidence: dict,
) -> tuple[str, str, str]:
    """Return (james_decision, james_notes, routing).

    Routing values (2026-05-09 evidence-aware revision):
      - 'pre_decided'           — James-pre-decided override applies.
      - 'auto_resolved'         — matched / structural-alias / productive-
                                  multiplicity per already-decided policy.
      - 'james_reviewable'      — HIGH/MEDIUM evidence, no conflict, not
                                  auto-resolvable. James can directly
                                  adjudicate without Red.
      - 'needs_red'             — HIGH/MEDIUM evidence + confirmed
                                  conflict signal.
      - 'provisional_candidate' — LOW evidence, no conflict. Hold for
                                  more evidence before promoting or
                                  asking Red.
      - 'quarantine'            — LOW evidence + conflict. Defer until
                                  evidence improves; not Red-worthy yet.

    Policy preserved:
      - Footbag.org rows are valid external evidence by default.
      - Red review reserved ONLY for genuine ambiguity/conflict at
        HIGH/MEDIUM evidence.
      - Missing-from-IFPA never alone implies low validity.
    """
    # 1. Pre-decided overrides.
    key = _normalize_name_key(name)
    pre_decided_norm = {_normalize_name_key(k): v for k, v in JAMES_PRE_DECIDED.items()}
    if key in pre_decided_norm:
        decision, note = pre_decided_norm[key]
        return (decision, note, "pre_decided")

    has_adds_conflict = (diff_row.get("adds_conflict") or "").lower().startswith("yes")
    has_status_conflict = recon_status in CONFIRMED_CONFLICT_STATUSES
    has_conflict = has_adds_conflict or has_status_conflict

    confidence = evidence["evidence_confidence"]

    # 2. Conflict path — split by evidence band.
    if has_conflict:
        if confidence == "LOW":
            return ("", "", "quarantine")
        # HIGH/MEDIUM evidence + confirmed conflict → Red.
        return ("", "", "needs_red")

    # 3. No-conflict, auto-resolvable categories (matched / structural / mult).
    if recon_status in AUTO_RESOLVED_STATUSES:
        return ("", "", "auto_resolved")

    # 4. No-conflict, no auto-resolution: split by evidence band.
    if confidence == "LOW":
        return ("", "", "provisional_candidate")
    return ("", "", "james_reviewable")


def build_workbook(
    diff_path: Path, adj_path: Path, db_path: Path, out_path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    db = sqlite3.connect(str(db_path))
    canonical     = load_canonical(db)
    trick_sources = load_trick_source_index(db)
    video_set     = load_trick_video_set(db)
    # R1 additions — load Tricks/Aliases data for the new sheets.
    trick_rows = load_trick_index_for_workbook(db)
    alias_rows = load_aliases_for_workbook(db)
    db.close()

    red_queue_rows = seed_red_queue()

    adjudication = load_adjudication(adj_path)

    with diff_path.open(newline="", encoding="utf-8") as f:
        diff_rows = list(csv.DictReader(f))

    # R1 — fborg cross-reference for the Tricks sheet. Index matched diff
    # rows by canonical_slug so each Tricks row can carry the externally-
    # asserted name + ADD without re-running the reconciliation pipeline.
    fborg_by_slug: dict[str, dict] = {}
    for d in diff_rows:
        if d.get("canonical_match") == "yes":
            cs = d.get("canonical_slug") or ""
            if cs and cs not in fborg_by_slug:
                fborg_by_slug[cs] = {
                    "fborg_name": d.get("external_name", ""),
                    "fborg_add":  parse_int(d.get("external_ADD", "")) or "",
                }

    # R1 — Red queue cross-reference: which slugs have a pending review row,
    # and what topic/id to surface on the corresponding Tricks row.
    red_lookup: dict[str, dict] = {}
    for q in red_queue_rows:
        slug = q.get("slug_affected") or ""
        if slug and slug not in red_lookup:
            red_lookup[slug] = {
                "red_review_topic": q.get("topic", ""),
                "red_review_id":    q.get("id", ""),
            }

    # Column layout. Source-presence columns are generated from SOURCES
    # so adding a future source extends the registry, not the list here.
    source_cols = [(f"evidence_{sid}", 8) for sid in SOURCE_IDS]
    columns = [
        # Footbag.org side
        ("showmove_id",          12),
        ("fborg_name",           28),
        ("fborg_url",            32),
        ("fborg_ADD",             6),
        ("fborg_notation",       40),
        ("fborg_description",    48),
        ("fborg_alt_name",       22),
        ("has_video",             8),
        # IFPA side
        ("ifpa_slug",            22),
        ("ifpa_name",            22),
        ("ifpa_ADD",              6),
        ("ifpa_notation",        40),
        ("ifpa_description",     40),
        ("ifpa_status",          12),
        # Reconciliation
        ("match_type",           18),
        ("reconciliation_status",30),
        # Evidence: per-source presence + derived signals
        *source_cols,
        ("evidence_jobs_notation",     8),
        ("evidence_structural_parse",  8),
        ("evidence_existing_ifpa",     8),
        ("evidence_video",             8),
        ("evidence_multiple_sources",  8),
        ("evidence_total_score",       6),
        ("evidence_confidence",       10),
        # Lineage
        ("first_seen_source",         12),
        ("strongest_source",          12),
        ("supporting_sources",        38),
        # Comment + adjudication
        ("comment",                   60),
        ("routing",                   22),
        ("james_decision",            38),
        ("james_notes",               48),
        ("red_status",                24),
    ]

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A4D70")
    align_top   = Alignment(vertical="top", wrap_text=True)

    # R1 — column-group banding on the Tricks sheet header row. Visual only;
    # no data dependency. Row stripes use a separate palette below.
    tricks_group_fills = {
        "identity":       PatternFill("solid", fgColor="EAEAEA"),
        "editorial":      PatternFill("solid", fgColor="F5F5F5"),
        "notation":       PatternFill("solid", fgColor="DCEEF7"),  # light blue
        "parser":         PatternFill("solid", fgColor="FFF6CC"),  # light yellow
        "media":          PatternFill("solid", fgColor="DDEED9"),  # light green
        "external":       PatternFill("solid", fgColor="FCE5CD"),  # peach
        "decision":       PatternFill("solid", fgColor="E5DBF0"),  # light purple
        "audit":          PatternFill("solid", fgColor="F5F5F5"),
    }
    # Map each TRICKS_COLUMNS index to its group (1-based).
    tricks_col_groups = (
        ["identity"]*6 + ["editorial"]*1 + ["notation"]*5 + ["parser"]*3 +
        ["media"]*3 + ["external"]*4 + ["decision"]*6 + ["audit"]*2
    )
    assert len(tricks_col_groups) == len(TRICKS_COLUMNS), (
        f"tricks_col_groups ({len(tricks_col_groups)}) must match "
        f"TRICKS_COLUMNS ({len(TRICKS_COLUMNS)})"
    )

    pending_fill        = PatternFill("solid", fgColor="FFF2CC")  # yellow — *_status = pending / red_review_needed
    notation_authored_fill = PatternFill("solid", fgColor="FFFFFF")
    advanced_fill       = PatternFill("solid", fgColor="E8F0F7")  # light blue — page_completeness = advanced
    showcase_fill       = PatternFill("solid", fgColor="D9EAD3")  # green — page_completeness = showcase
    minimal_fill        = PatternFill("solid", fgColor="F5F5F5")  # gray — page_completeness = minimal

    # ─── Sheet 1: Tricks ────────────────────────────────────────────────────
    tricks_ws = wb.active
    tricks_ws.title = "Tricks"
    for idx, (name, width) in enumerate(TRICKS_COLUMNS, start=1):
        c = tricks_ws.cell(row=1, column=idx, value=name)
        c.font = header_font
        c.fill = tricks_group_fills[tricks_col_groups[idx - 1]]
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        tricks_ws.column_dimensions[get_column_letter(idx)].width = width
    tricks_ws.row_dimensions[1].height = 30
    tricks_ws.freeze_panes = "C2"  # keep slug + display_name visible while scrolling

    tricks_summary = {b: 0 for b in PAGE_COMPLETENESS_BANDS}
    notation_status_counts: dict[str, int] = {}
    tutorial_status_counts: dict[str, int] = {}
    parser_status_counts:   dict[str, int] = {}
    adoption_status_counts: dict[str, int] = {}

    for tr in trick_rows:
        # Cross-references discovered above
        slug = tr["slug"]
        if slug in fborg_by_slug:
            tr["fborg_match"] = "matched"
            tr["fborg_name"]  = fborg_by_slug[slug]["fborg_name"]
            tr["fborg_add"]   = fborg_by_slug[slug]["fborg_add"]
        else:
            tr["fborg_match"] = "no_external_match"
        if slug in red_lookup:
            tr["red_review_needed"] = "TRUE"
            tr["red_review_topic"]  = red_lookup[slug]["red_review_topic"]
            tr["red_review_id"]     = red_lookup[slug]["red_review_id"]

        target = tricks_ws.max_row + 1
        for col_idx, (name, _) in enumerate(TRICKS_COLUMNS, start=1):
            c = tricks_ws.cell(row=target, column=col_idx, value=tr.get(name, ""))
            c.alignment = align_top

        # Visual cues:
        # 1. Whole-row tint by page_completeness band (most important reader signal).
        # 2. Yellow on red_review_needed cell (focus filter).
        # 3. Yellow on semantic_notation_status when 'pending'.
        band = tr["page_completeness"]
        if band == "showcase":
            row_fill = showcase_fill
        elif band == "advanced":
            row_fill = advanced_fill
        elif band == "minimal":
            row_fill = minimal_fill
        else:
            row_fill = None
        if row_fill is not None:
            for col_idx in range(1, len(TRICKS_COLUMNS) + 1):
                tricks_ws.cell(row=target, column=col_idx).fill = row_fill
        # Focused cell tints (override row tint where attention is warranted).
        if tr["red_review_needed"] == "TRUE":
            col_idx = [c[0] for c in TRICKS_COLUMNS].index("red_review_needed") + 1
            tricks_ws.cell(row=target, column=col_idx).fill = pending_fill
        if tr["semantic_notation_status"] == "pending":
            col_idx = [c[0] for c in TRICKS_COLUMNS].index("semantic_notation_status") + 1
            tricks_ws.cell(row=target, column=col_idx).fill = pending_fill

        tricks_summary[band] += 1
        notation_status_counts[tr["semantic_notation_status"]] = notation_status_counts.get(tr["semantic_notation_status"], 0) + 1
        tutorial_status_counts[tr["tutorial_status"]]          = tutorial_status_counts.get(tr["tutorial_status"], 0) + 1
        parser_status_counts[tr["parser_status"]]              = parser_status_counts.get(tr["parser_status"], 0) + 1
        adoption_status_counts[tr["adoption_status"]]          = adoption_status_counts.get(tr["adoption_status"], 0) + 1

    tricks_ws.auto_filter.ref = f"A1:{get_column_letter(len(TRICKS_COLUMNS))}{tricks_ws.max_row}"

    # ─── Sheet 2: Reconciliation (existing logic, additive — sheet ordering only) ─
    ws = wb.create_sheet("Reconciliation")

    james_review_fill  = PatternFill("solid", fgColor="FFF2CC")  # yellow
    needs_red_fill     = PatternFill("solid", fgColor="F4CCCC")  # red
    pre_decided_fill   = PatternFill("solid", fgColor="D9EAD3")  # green
    matched_fill       = PatternFill("solid", fgColor="E8F0F7")  # light blue
    auto_resolved_fill = PatternFill("solid", fgColor="EFEFEF")  # neutral gray
    provisional_fill   = PatternFill("solid", fgColor="FCE5CD")  # peach (LOW + no conflict)
    quarantine_fill    = PatternFill("solid", fgColor="D5A6BD")  # mauve (LOW + conflict)

    for idx, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row=1, column=idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    summary = {
        "auto_resolved":         0,
        "james_reviewable":      0,
        "needs_red":             0,
        "pre_decided":           0,
        "provisional_candidate": 0,
        "quarantine":            0,
    }
    # Cross-tab counters for the Evidence Summary sheet.
    confidence_counts:  dict[str, int] = {c: 0 for c in CONFIDENCE_LABELS}
    strongest_counts:   dict[str, int] = {sid: 0 for sid in SOURCE_IDS}
    strongest_counts[""] = 0   # row with no sources present (defensive)
    routing_x_confidence: dict[tuple[str, str], int] = {}
    overlap_counts:     dict[str, int] = {}
    conflict_counts = {"none": 0, "adds_conflict": 0, "policy_blocked_down": 0,
                       "canonical_candidate": 0, "productive_mult_base_unresolved": 0,
                       "structural_alias_add_conflict": 0, "other_conflict": 0}

    for diff in diff_rows:
        external_name = diff.get("external_name", "")
        external_slug = diff.get("external_slug") or slugify(external_name)
        canonical_slug = diff.get("canonical_slug") or ""
        canonical_match = (diff.get("canonical_match") == "yes")

        ifpa_row = canonical.get(canonical_slug) if canonical_match else None
        ifpa_status = (
            ("active" if ifpa_row["is_active"] else "pending")
            if ifpa_row else "not_in_dict"
        )

        adj_row = adjudication.get(external_slug)
        recon_status, comment = derive_recon_status(diff, adj_row)

        evidence = evaluate_evidence(diff, adj_row, ifpa_row, trick_sources, video_set)

        james_decision, james_notes, routing = adjudication_for(
            external_name, recon_status, diff, adj_row, evidence
        )
        red_status = (adj_row or {}).get("auto_classification", "")

        # Provisional slug for unmatched rows (fborg_supported_missing etc.).
        ifpa_slug_out = canonical_slug if canonical_match else f"(provisional) {external_slug}"

        row_values = [
            diff.get("external_showmove_id", ""),
            external_name,
            diff.get("external_url", ""),
            parse_int(diff.get("external_ADD", "")) or "",
            diff.get("external_notation", ""),
            diff.get("external_description", ""),
            diff.get("external_alt_name", ""),
            "yes" if diff.get("media_present") == "yes" or diff.get("has_media") == "yes" else "no",
            ifpa_slug_out,
            (ifpa_row or {}).get("name", ""),
            (ifpa_row or {}).get("adds") if ifpa_row else "",
            (ifpa_row or {}).get("notation", ""),
            (ifpa_row or {}).get("description", ""),
            ifpa_status,
            diff.get("match_type", ""),
            recon_status,
            # Evidence presence (per-source in registry order)
            *(evidence[f"evidence_{sid}"] for sid in SOURCE_IDS),
            # Evidence: derived signals
            evidence["evidence_jobs_notation"],
            evidence["evidence_structural_parse"],
            evidence["evidence_existing_ifpa"],
            evidence["evidence_video"],
            evidence["evidence_multiple_sources"],
            evidence["evidence_total_score"],
            evidence["evidence_confidence"],
            # Lineage
            evidence["first_seen_source"],
            evidence["strongest_source"],
            evidence["supporting_sources"],
            # Comment + adjudication
            comment,
            routing,
            james_decision,
            james_notes,
            red_status,
        ]

        target_row = ws.max_row + 1
        for col_idx, val in enumerate(row_values, start=1):
            c = ws.cell(row=target_row, column=col_idx, value=val)
            c.alignment = align_top

        # Whole-row highlight by routing.
        if routing == "pre_decided":
            fill = pre_decided_fill
        elif routing == "needs_red":
            fill = needs_red_fill
        elif routing == "james_reviewable":
            fill = james_review_fill
        elif routing == "provisional_candidate":
            fill = provisional_fill
        elif routing == "quarantine":
            fill = quarantine_fill
        elif routing == "auto_resolved" and canonical_match:
            fill = matched_fill
        elif routing == "auto_resolved":
            fill = auto_resolved_fill
        else:
            fill = None
        summary[routing] = summary.get(routing, 0) + 1
        if fill is not None:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=target_row, column=col_idx).fill = fill

        # Cross-tab accumulators for the Evidence Summary sheet.
        confidence_counts[evidence["evidence_confidence"]] = confidence_counts.get(evidence["evidence_confidence"], 0) + 1
        strongest_counts[evidence["strongest_source"]] = strongest_counts.get(evidence["strongest_source"], 0) + 1
        routing_x_confidence[(routing, evidence["evidence_confidence"])] = (
            routing_x_confidence.get((routing, evidence["evidence_confidence"]), 0) + 1
        )
        # Conflict-type accumulator.
        if (diff.get("adds_conflict") or "").lower().startswith("yes"):
            conflict_counts["adds_conflict"] += 1
        elif recon_status == "policy_blocked_down":
            conflict_counts["policy_blocked_down"] += 1
        elif recon_status == "canonical_candidate":
            conflict_counts["canonical_candidate"] += 1
        elif recon_status == "productive_multiplicity_base_unresolved":
            conflict_counts["productive_mult_base_unresolved"] += 1
        elif recon_status == "add_conflict":
            conflict_counts["structural_alias_add_conflict"] += 1
        elif recon_status in CONFIRMED_CONFLICT_STATUSES:
            conflict_counts["other_conflict"] += 1
        else:
            conflict_counts["none"] += 1
        # Source-overlap signature (sorted, comma-joined).
        overlap_counts[evidence["supporting_sources"]] = (
            overlap_counts.get(evidence["supporting_sources"], 0) + 1
        )

    # ─── Sheet 3: Aliases ───────────────────────────────────────────────────
    aliases_ws = wb.create_sheet("Aliases")
    for idx, (name, width) in enumerate(ALIASES_COLUMNS, start=1):
        c = aliases_ws.cell(row=1, column=idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        aliases_ws.column_dimensions[get_column_letter(idx)].width = width
    aliases_ws.row_dimensions[1].height = 22
    aliases_ws.freeze_panes = "A2"
    for ar in alias_rows:
        target = aliases_ws.max_row + 1
        for col_idx, (name, _) in enumerate(ALIASES_COLUMNS, start=1):
            c = aliases_ws.cell(row=target, column=col_idx, value=ar.get(name, ""))
            c.alignment = align_top
    if alias_rows:
        aliases_ws.auto_filter.ref = f"A1:{get_column_letter(len(ALIASES_COLUMNS))}{aliases_ws.max_row}"

    # ─── Sheet 4: Red Review Queue ──────────────────────────────────────────
    rrq_ws = wb.create_sheet("Red Review Queue")
    for idx, (name, width) in enumerate(RED_QUEUE_COLUMNS, start=1):
        c = rrq_ws.cell(row=1, column=idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        rrq_ws.column_dimensions[get_column_letter(idx)].width = width
    rrq_ws.row_dimensions[1].height = 22
    rrq_ws.freeze_panes = "B2"
    for q in red_queue_rows:
        target = rrq_ws.max_row + 1
        for col_idx, (name, _) in enumerate(RED_QUEUE_COLUMNS, start=1):
            c = rrq_ws.cell(row=target, column=col_idx, value=q.get(name, ""))
            c.alignment = align_top
        # Yellow tint on row when status is pending.
        if q.get("red_status") == "pending":
            for col_idx in range(1, len(RED_QUEUE_COLUMNS) + 1):
                rrq_ws.cell(row=target, column=col_idx).fill = pending_fill
    if red_queue_rows:
        rrq_ws.auto_filter.ref = f"A1:{get_column_letter(len(RED_QUEUE_COLUMNS))}{rrq_ws.max_row}"

    # Summary sheet.
    s = wb.create_sheet("Summary")
    s["A1"] = "Footbag.org × IFPA reconciliation — counts"
    s["A1"].font = Font(bold=True, size=14)
    s["A3"] = "Total rows"; s["B3"] = len(diff_rows)
    s["A4"] = "Auto-resolved (matched / structural-alias / productive-multiplicity)"
    s["B4"] = summary["auto_resolved"]
    s["A5"] = "James-reviewable (HIGH/MEDIUM evidence, no conflict)"
    s["B5"] = summary["james_reviewable"]
    s["A6"] = "Provisional candidate (LOW evidence, no conflict)"
    s["B6"] = summary["provisional_candidate"]
    s["A7"] = "James pre-decided"
    s["B7"] = summary["pre_decided"]
    s["A8"] = "Needs Red review (HIGH/MEDIUM evidence + confirmed conflict)"
    s["B8"] = summary["needs_red"]
    s["A9"] = "Quarantine (LOW evidence + conflict)"
    s["B9"] = summary["quarantine"]
    s["A10"] = "Unrouted (defensive — should be 0)"
    s["B10"] = len(diff_rows) - sum(summary.values())

    s["A12"] = "Routing policy (evidence-aware, 2026-05-09):"; s["A12"].font = Font(bold=True)
    s["A13"] = "Footbag.org rows are valid external evidence by default. A row's routing depends on (confidence band, conflict signal):"
    s["A14"] = "  HIGH/MEDIUM + no conflict   → auto_resolved | james_reviewable"
    s["A15"] = "  HIGH/MEDIUM + conflict      → needs_red"
    s["A16"] = "  LOW         + no conflict   → provisional_candidate"
    s["A17"] = "  LOW         + conflict      → quarantine"
    s["A18"] = "Confirmed conflict signals: ADD math conflict, down-family policy deferral, productive-multiplicity base unresolved, canonical-elevation question."

    s["A20"] = "Color legend (Reconciliation sheet rows):"; s["A20"].font = Font(bold=True)
    s["A21"] = "Light blue: matched canonical (auto_resolved)"; s["A21"].fill = matched_fill
    s["A22"] = "Gray: structural-alias / productive-multiplicity (auto_resolved)"; s["A22"].fill = auto_resolved_fill
    s["A23"] = "Yellow: james_reviewable (HIGH/MEDIUM evidence, no conflict)"; s["A23"].fill = james_review_fill
    s["A24"] = "Peach: provisional_candidate (LOW evidence, no conflict)"; s["A24"].fill = provisional_fill
    s["A25"] = "Green: James pre-decided"; s["A25"].fill = pre_decided_fill
    s["A26"] = "Red: needs_red (HIGH/MEDIUM evidence + conflict)"; s["A26"].fill = needs_red_fill
    s["A27"] = "Mauve: quarantine (LOW evidence + conflict)"; s["A27"].fill = quarantine_fill

    # R1 — Tricks-sheet aggregates (new section appended below the existing
    # Reconciliation legend so the original counts stay where curators
    # already know to look).
    s["A29"] = "Tricks sheet — aggregates (R1)"; s["A29"].font = Font(bold=True, size=12)
    s["A30"] = "Total active tricks";                                s["B30"] = len(trick_rows)
    row = 32
    s.cell(row=row, column=1, value="Page completeness band").font = Font(bold=True)
    s.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for band in PAGE_COMPLETENESS_BANDS:
        s.cell(row=row, column=1, value=band)
        s.cell(row=row, column=2, value=tricks_summary.get(band, 0))
        row += 1
    row += 1
    s.cell(row=row, column=1, value="Semantic notation status").font = Font(bold=True)
    s.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for stat in ("authored", "parser-derived", "pending"):
        s.cell(row=row, column=1, value=stat)
        s.cell(row=row, column=2, value=notation_status_counts.get(stat, 0))
        row += 1
    row += 1
    s.cell(row=row, column=1, value="Tutorial status").font = Font(bold=True)
    s.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for stat in ("tutorial-coverage", "demo-only", "no-video"):
        s.cell(row=row, column=1, value=stat)
        s.cell(row=row, column=2, value=tutorial_status_counts.get(stat, 0))
        row += 1
    row += 1
    s.cell(row=row, column=1, value="Parser status").font = Font(bold=True)
    s.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for stat in ("exact_modifier_derived", "exact_self_atom", "approximate", "policy_dependent", "unresolved", "not-run"):
        n = parser_status_counts.get(stat, 0)
        if n == 0 and stat in ("approximate", "policy_dependent", "unresolved", "not-run"):
            continue
        s.cell(row=row, column=1, value=stat)
        s.cell(row=row, column=2, value=n)
        row += 1
    row += 1
    s.cell(row=row, column=1, value="Adoption status").font = Font(bold=True)
    s.cell(row=row, column=2, value="Count").font = Font(bold=True)
    row += 1
    for stat in ("live", "draft", "hidden", "deprecated"):
        s.cell(row=row, column=1, value=stat)
        s.cell(row=row, column=2, value=adoption_status_counts.get(stat, 0))
        row += 1
    row += 1
    s.cell(row=row, column=1, value="Red review queue size (R1 seed)").font = Font(bold=True)
    s.cell(row=row, column=2, value=len(red_queue_rows))

    s.column_dimensions["A"].width = 80
    s.column_dimensions["B"].width = 12

    # ─── Evidence Summary sheet ─────────────────────────────────────────────
    es = wb.create_sheet("Evidence Summary")
    es["A1"] = "Evidence-density and lineage distributions"
    es["A1"].font = Font(bold=True, size=14)

    row = 3
    es.cell(row=row, column=1, value="Counts by confidence band").font = Font(bold=True); row += 1
    es.cell(row=row, column=1, value="Confidence"); es.cell(row=row, column=2, value="Count")
    es.cell(row=row, column=1).font = Font(bold=True); es.cell(row=row, column=2).font = Font(bold=True); row += 1
    for c in ("HIGH", "MEDIUM", "LOW"):
        es.cell(row=row, column=1, value=c)
        es.cell(row=row, column=2, value=confidence_counts.get(c, 0))
        row += 1

    row += 1
    es.cell(row=row, column=1, value="Counts by strongest_source").font = Font(bold=True); row += 1
    es.cell(row=row, column=1, value="Source"); es.cell(row=row, column=2, value="Count")
    es.cell(row=row, column=1).font = Font(bold=True); es.cell(row=row, column=2).font = Font(bold=True); row += 1
    # Sort by registry strength (highest first), then count desc.
    for sid in sorted(SOURCE_IDS, key=lambda s: -SOURCE_STRENGTH[s]):
        n = strongest_counts.get(sid, 0)
        if n == 0: continue
        es.cell(row=row, column=1, value=f"{sid} — {SOURCE_LABEL[sid]}")
        es.cell(row=row, column=2, value=n)
        row += 1
    if strongest_counts.get("", 0):
        es.cell(row=row, column=1, value="(no source — defensive)")
        es.cell(row=row, column=2, value=strongest_counts[""])
        row += 1

    row += 1
    es.cell(row=row, column=1, value="Counts by routing").font = Font(bold=True); row += 1
    es.cell(row=row, column=1, value="Routing"); es.cell(row=row, column=2, value="Count")
    es.cell(row=row, column=1).font = Font(bold=True); es.cell(row=row, column=2).font = Font(bold=True); row += 1
    for r in ("auto_resolved", "james_reviewable", "provisional_candidate",
              "pre_decided", "needs_red", "quarantine"):
        es.cell(row=row, column=1, value=r)
        es.cell(row=row, column=2, value=summary.get(r, 0))
        row += 1

    row += 1
    es.cell(row=row, column=1, value="Counts by conflict type").font = Font(bold=True); row += 1
    es.cell(row=row, column=1, value="Conflict type"); es.cell(row=row, column=2, value="Count")
    es.cell(row=row, column=1).font = Font(bold=True); es.cell(row=row, column=2).font = Font(bold=True); row += 1
    for k, v in sorted(conflict_counts.items(), key=lambda kv: (-kv[1], kv[0])):
        es.cell(row=row, column=1, value=k)
        es.cell(row=row, column=2, value=v)
        row += 1

    row += 1
    es.cell(row=row, column=1, value="Routing × confidence cross-tab").font = Font(bold=True); row += 1
    es.cell(row=row, column=1, value="Routing"); es.cell(row=row, column=2, value="Confidence"); es.cell(row=row, column=3, value="Count")
    for col in (1, 2, 3): es.cell(row=row, column=col).font = Font(bold=True)
    row += 1
    for (rt, conf), n in sorted(routing_x_confidence.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        es.cell(row=row, column=1, value=rt)
        es.cell(row=row, column=2, value=conf)
        es.cell(row=row, column=3, value=n)
        row += 1

    row += 1
    es.cell(row=row, column=1, value="Top source-overlap signatures (combinations of sources present)").font = Font(bold=True); row += 1
    es.cell(row=row, column=1, value="Sources present"); es.cell(row=row, column=2, value="Count")
    es.cell(row=row, column=1).font = Font(bold=True); es.cell(row=row, column=2).font = Font(bold=True); row += 1
    for sig, n in sorted(overlap_counts.items(), key=lambda kv: -kv[1])[:25]:
        es.cell(row=row, column=1, value=sig or "(none)")
        es.cell(row=row, column=2, value=n)
        row += 1

    es.column_dimensions["A"].width = 60
    es.column_dimensions["B"].width = 14
    es.column_dimensions["C"].width = 14

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Wrote: {out_path}")
    print("By routing:")
    print(f"  auto_resolved:         {summary['auto_resolved']}")
    print(f"  james_reviewable:      {summary['james_reviewable']}")
    print(f"  provisional_candidate: {summary['provisional_candidate']}")
    print(f"  james_pre_decided:     {summary['pre_decided']}")
    print(f"  needs_red:             {summary['needs_red']}")
    print(f"  quarantine:            {summary['quarantine']}")
    unrouted = len(diff_rows) - sum(summary.values())
    if unrouted: print(f"  UNROUTED (bug?):       {unrouted}")
    print(f"  total:                 {len(diff_rows)}")
    print("By confidence:")
    for c in ("HIGH", "MEDIUM", "LOW"):
        print(f"  {c:<7s}              {confidence_counts.get(c, 0)}")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--diff",          default=str(DEFAULT_DIFF))
    ap.add_argument("--adjudication",  default=str(DEFAULT_ADJUDICATION))
    ap.add_argument("--db",            default=str(DEFAULT_DB))
    ap.add_argument("--out",           default=str(DEFAULT_OUT))
    args = ap.parse_args()

    for label, p in [("diff", args.diff), ("adjudication", args.adjudication), ("db", args.db)]:
        if not Path(p).exists():
            print(f"ERROR: {label} not found at {p}", file=sys.stderr)
            return 1

    build_workbook(Path(args.diff), Path(args.adjudication), Path(args.db), Path(args.out))
    return 0


if __name__ == "__main__":
    sys.exit(main())
