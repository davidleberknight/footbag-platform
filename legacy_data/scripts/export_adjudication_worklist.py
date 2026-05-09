"""
Export the james_reviewable rows from the reconciliation workbook into
a structured adjudication CSV that James edits to drive the importer.

Read-only with respect to ontology data. The output CSV is the only
write target.

Inputs:
  legacy_data/reports/fborg_reconciliation.xlsx     (workbook)
  legacy_data/freestyle_reconciliation/james_adjudications.csv (existing CSV, if any)

Output:
  legacy_data/freestyle_reconciliation/james_adjudications.csv

Idempotent merge: when the CSV already exists, decisions James has
filled in are preserved row-by-row keyed on `showmove_id`. The export
adds new james_reviewable rows from the workbook and refreshes the
read-only context columns (evidence/lineage/provenance) without
clobbering any decision payload column.

CSV schema:
  showmove_id              — link back to the workbook row
  fborg_name               — read-only context
  fborg_url                — read-only context (provenance)
  fborg_ADD                — read-only context
  fborg_notation           — read-only context
  evidence_score           — read-only context
  evidence_confidence      — read-only context
  strongest_source         — read-only context
  supporting_sources       — read-only context

  decision                 — James-edited: ADD_CANONICAL | ADD_ALIAS |
                             PRODUCTIVE_VARIANT | PROVISIONAL | REJECT
                             (blank = pending)
  target_canonical_slug    — for ADD_CANONICAL or ADD_ALIAS target
  target_canonical_name    — for ADD_CANONICAL
  target_canonical_adds    — for ADD_CANONICAL
  target_canonical_notation — for ADD_CANONICAL
  target_canonical_description — for ADD_CANONICAL
  target_canonical_family  — for ADD_CANONICAL
  target_canonical_base_trick — for ADD_CANONICAL
  alias_text               — for ADD_ALIAS
  alias_type               — for ADD_ALIAS (default 'common')
  notes                    — free text rationale

Does not modify any ontology data, DB, or other CSVs.
"""
import argparse
import csv
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = REPO / "legacy_data" / "reports" / "fborg_reconciliation.xlsx"
DEFAULT_CSV  = REPO / "legacy_data" / "freestyle_reconciliation" / "james_adjudications.csv"

# Ordered column list. Read-only context first, then decision payload.
CONTEXT_COLUMNS = [
    "showmove_id",
    "fborg_name",
    "fborg_url",
    "fborg_ADD",
    "fborg_notation",
    "evidence_score",
    "evidence_confidence",
    "strongest_source",
    "supporting_sources",
]
DECISION_COLUMNS = [
    "decision",
    "target_canonical_slug",
    "target_canonical_name",
    "target_canonical_adds",
    "target_canonical_notation",
    "target_canonical_description",
    "target_canonical_family",
    "target_canonical_base_trick",
    "alias_text",
    "alias_type",
    "notes",
]
ALL_COLUMNS = CONTEXT_COLUMNS + DECISION_COLUMNS

DEFAULT_FILTER_ROUTING = {"james_reviewable"}


def load_existing_csv(path: Path) -> dict[str, dict]:
    """Read existing adjudications keyed by showmove_id; preserves the
    decision-payload columns so re-export doesn't clobber James's edits."""
    if not path.exists():
        return {}
    out = {}
    with path.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            sid = (row.get("showmove_id") or "").strip()
            if sid:
                out[sid] = row
    return out


def load_workbook_rows(xlsx_path: Path, filter_routing: set[str]) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Reconciliation" not in wb.sheetnames:
        raise SystemExit(f"ERROR: 'Reconciliation' sheet not found in {xlsx_path}")
    ws = wb["Reconciliation"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    col = {name: i for i, name in enumerate(header)}
    required = {
        "showmove_id", "fborg_name", "fborg_url", "fborg_ADD", "fborg_notation",
        "evidence_total_score", "evidence_confidence",
        "strongest_source", "supporting_sources", "routing",
    }
    missing = required - col.keys()
    if missing:
        raise SystemExit(f"ERROR: workbook missing columns: {sorted(missing)}")
    out = []
    for r in rows[1:]:
        routing = r[col["routing"]]
        if routing not in filter_routing:
            continue
        out.append({
            "showmove_id":          str(r[col["showmove_id"]] or "").strip(),
            "fborg_name":           r[col["fborg_name"]] or "",
            "fborg_url":            r[col["fborg_url"]] or "",
            "fborg_ADD":            r[col["fborg_ADD"]] if r[col["fborg_ADD"]] is not None else "",
            "fborg_notation":       r[col["fborg_notation"]] or "",
            "evidence_score":       r[col["evidence_total_score"]] if r[col["evidence_total_score"]] is not None else "",
            "evidence_confidence":  r[col["evidence_confidence"]] or "",
            "strongest_source":     r[col["strongest_source"]] or "",
            "supporting_sources":   r[col["supporting_sources"]] or "",
        })
    return out


def merge_rows(workbook_rows: list[dict], existing: dict[str, dict]) -> list[dict]:
    """For each workbook row, refresh context columns; preserve decision
    columns from `existing` if present. Rows in `existing` whose
    showmove_id is not in `workbook_rows` are kept (so the operator's
    in-progress decisions for rows that may have shifted routing aren't
    silently dropped — surfaced for review)."""
    merged = []
    seen = set()
    for w in workbook_rows:
        sid = w["showmove_id"]
        seen.add(sid)
        prev = existing.get(sid, {})
        row = {k: w.get(k, "") for k in CONTEXT_COLUMNS}
        for c in DECISION_COLUMNS:
            row[c] = prev.get(c, "")
        merged.append(row)
    # Carry forward orphaned existing rows (defensive).
    for sid, prev in existing.items():
        if sid in seen:
            continue
        if any((prev.get(c) or "").strip() for c in DECISION_COLUMNS):
            row = {k: prev.get(k, "") for k in ALL_COLUMNS}
            merged.append(row)
    return merged


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--csv",  default=str(DEFAULT_CSV))
    ap.add_argument("--routing", action="append",
                    help="Routing values to include (default: james_reviewable). Repeatable.")
    args = ap.parse_args()

    routing = set(args.routing) if args.routing else DEFAULT_FILTER_ROUTING
    xlsx = Path(args.xlsx); csv_path = Path(args.csv)
    if not xlsx.exists():
        print(f"ERROR: workbook not found at {xlsx}", file=sys.stderr)
        return 1
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    workbook_rows = load_workbook_rows(xlsx, routing)
    existing      = load_existing_csv(csv_path)
    merged        = merge_rows(workbook_rows, existing)

    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=ALL_COLUMNS)
        w.writeheader()
        w.writerows(merged)

    n_with_decision = sum(1 for r in merged if (r.get("decision") or "").strip())
    print(f"Wrote: {csv_path}")
    print(f"  routing filter: {sorted(routing)}")
    print(f"  rows from workbook: {len(workbook_rows)}")
    print(f"  preserved from prior CSV: {len(existing)}")
    print(f"  merged total: {len(merged)}")
    print(f"  rows with a decision filled in: {n_with_decision}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
