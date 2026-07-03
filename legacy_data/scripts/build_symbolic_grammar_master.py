#!/usr/bin/env python3
"""SYMBOLIC-GRAMMAR-1: build master FootbagMoves/PassBack symbolic spreadsheet.

OBSERVATIONAL symbolic-grammar layer; independent of IFPA canonical dictionary.

Reads:
- legacy_data/out/footbagmoves_inventory.csv  (573 rows; 377 with operational notation)
- freestyle/tools/trick_video_discovery/passback_review_queue.csv (106 rows; alternate-notation style)

Writes:
- exploration/footbagmoves-federation/SYMBOLIC_GRAMMAR_MASTER.xlsx
- freestyle/inputs/observational/SYMBOLIC_GRAMMAR_MASTER.csv
  (a live input of the freestyle content generators, so it lives in the
  self-contained freestyle tree)
- exploration/footbagmoves-federation/SYMBOLIC_GRAMMAR_UNRESOLVED_TOKENS.csv
- exploration/footbagmoves-federation/SYMBOLIC_GRAMMAR_FAMILY_SUMMARY.csv

Constraints (per task brief):
- NO mutation of IFPA dictionary tables.
- NO alias insertion.
- NO ADD-value mutation; source_adds preserved as recorded.
- derived_adds is diagnostic-only (additive heuristic).
- Symbolic families are observational topology, NOT IFPA trick_family.

Parser layered after OPERATIONAL_NOTATION_GRAMMAR.md F0 reconnaissance.
"""
from __future__ import annotations

import csv
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

ROOT = Path(__file__).resolve().parent.parent.parent
FM_INVENTORY = ROOT / "legacy_data" / "out" / "footbagmoves_inventory.csv"
PASSBACK_QUEUE = ROOT / "freestyle" / "tools" / "trick_video_discovery" / "passback_review_queue.csv"
OUT_DIR = ROOT / "exploration" / "footbagmoves-federation"
MASTER_XLSX = OUT_DIR / "SYMBOLIC_GRAMMAR_MASTER.xlsx"
# The CSV twin is read by the freestyle content generators, so the maintained
# copy lives under freestyle/inputs/ (self-contained freestyle tree).
MASTER_CSV = ROOT / "freestyle" / "inputs" / "observational" / "SYMBOLIC_GRAMMAR_MASTER.csv"
UNRESOLVED_CSV = OUT_DIR / "SYMBOLIC_GRAMMAR_UNRESOLVED_TOKENS.csv"
FAMILY_CSV = OUT_DIR / "SYMBOLIC_GRAMMAR_FAMILY_SUMMARY.csv"

# Output schema (per task brief §2). Derived-only fields are prefixed `derived_*`
# or explicitly named with `parsed_` / `_flags` to mark non-source provenance.
FIELDS = [
    "source",                       # 'footbagmoves' | 'passback'
    "source_file",                  # original paste-file or queue origin
    "move_name",                    # display/colloquial name
    "alternate_names",              # pipe-delimited
    "technical_name",               # structural/canonical reading from source
    "source_adds",                  # recorded ADD count (authoritative for the source)
    "symbolic_notation_raw",        # raw notation string (preserved verbatim)
    "start_contact",                # parser: first contact token (toe/clip/set/...)
    "uptime_dex_1",                 # parser: first uptime dex token
    "uptime_dex_2",                 # parser: second uptime dex token (if present)
    "uptime_state_flags",           # parser: pre-state flags during uptime (rooted/no-plant/back/front)
    "midtime_body_actions",         # parser: body-action tokens (spin/duck/dive)
    "downtime_dexes",               # parser: dex tokens immediately preceding the final contact
    "final_contact",                # parser: terminating contact (toe/clip/flapper/...)
    "surface_flags",                # parser: surface qualifiers (UNS unusual surface)
    "positional_flags",             # parser: positional qualifiers (XBD cross-body, etc.)
    "component_flags",              # parser: component-flag stack (DEX/BOD/PDX/...)
    "parsed_symbol_sequence",       # parser: normalized beat-by-beat sequence (debug)
    "derived_adds",                 # parser: additive-heuristic ADD estimate (diagnostic)
    "source_vs_derived_delta",      # parser: source_adds - derived_adds (signed)
    "derived_symbolic_family",      # parser: topology family slug (NOT IFPA family)
    "derived_topology_family",      # parser: broader topology bucket
    "unresolved_tokens",            # parser: pipe-delimited unrecognized tokens
    "parse_confidence",             # parser: high/medium/low/none
    "review_status",                # parser: ok / needs_review / source_has_no_notation
    "notes",                        # parser: provenance/flag notes
]

# ─────────────────────────────────────────────────────────────────────────
# Token inventory (per OPERATIONAL_NOTATION_GRAMMAR.md F0 reconnaissance)
# ─────────────────────────────────────────────────────────────────────────

# Component flags (parens or brackets, ALL-CAPS post-token). +1 ADD contributors.
COMPONENT_FLAGS_ADD = {"DEX": 1, "DEL": 1, "BOD": 1, "XBD": 1, "PDX": 1, "XDEX": 1}
COMPONENT_FLAGS_NO_ADD = {"UNS"}  # unusual surface; not ADD-additive in heuristic
ALL_COMPONENT_FLAGS = set(COMPONENT_FLAGS_ADD.keys()) | COMPONENT_FLAGS_NO_ADD

# Pre-move state flags (parens lowercase, pre-token). Modify the IMMEDIATELY-FOLLOWING token.
PRE_STATE_FLAGS = {"back", "front", "no plant while", "rooted"}

# Contact/surface tokens (initial OR terminal). Used as start/end markers.
CONTACT_TOKENS = {"toe", "clip", "set", "flapper", "inside", "outside", "rake"}

# Side prefixes
SIDE_PREFIXES = {"same", "op"}

# Direction tokens (after side prefix, before contact)
DIRECTION_TOKENS = {"in", "out", "front", "back"}

# Body actions (verb-form, often after `(back)`/`(front)`)
BODY_ACTIONS = {"spin", "duck", "dive"}

# Whirl/swirl execution variants (compound multi-word, but observed as bigrams)
WHIRL_SWIRL_VARIANTS = {"front whirl", "back whirl", "front swirl", "back swirl"}

# Sequence operators
BEAT_GAP = ">>"
BEAT_FLOW = ">"


# ─────────────────────────────────────────────────────────────────────────
# Parsing dataclasses
# ─────────────────────────────────────────────────────────────────────────

@dataclass
class Beat:
    """One beat-segment of a notation string (between `>` or `>>` separators)."""
    raw: str
    pre_state: list[str] = field(default_factory=list)       # back / no plant while / rooted
    side: str | None = None                                  # same / op
    direction: str | None = None                             # in / out / front / back
    base_token: str | None = None                            # toe / clip / spin / duck / whirl / swirl
    qualifier: str | None = None                             # front-whirl, back-swirl, etc
    component_flags: list[str] = field(default_factory=list) # DEX, DEL, XBD, BOD, PDX, XDEX, UNS
    unresolved: list[str] = field(default_factory=list)
    is_gap_after: bool = False                               # `>>` followed this beat


@dataclass
class ParseResult:
    beats: list[Beat] = field(default_factory=list)
    unresolved_tokens: list[str] = field(default_factory=list)
    confidence: str = "none"

    @property
    def all_component_flags(self) -> list[str]:
        return [f for b in self.beats for f in b.component_flags]


# ─────────────────────────────────────────────────────────────────────────
# Notation normalization + tokenizer
# ─────────────────────────────────────────────────────────────────────────

def normalize_notation(raw: str) -> str:
    """Normalize notation: bracket→paren, case-fold structural tokens, collapse whitespace."""
    if not raw:
        return ""
    s = raw.strip()
    # Convert [X] component flags to (X) (only when the bracket content is ALL-CAPS short token)
    s = re.sub(r"\[([A-Z]{2,5})\]", r"(\1)", s)
    # Title-case "or" connector stays lowercase; uppercase any-case structural tokens to title
    # (We'll do case-insensitive token matching downstream, so just collapse extra whitespace.)
    s = re.sub(r"\s+", " ", s)
    # Preserve `>` and `>>` boundaries with surrounding spaces
    s = re.sub(r"\s*>>\s*", " >> ", s)
    s = re.sub(r"(?<![>])\s*>\s*(?![>])", " > ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


_BEAT_SPLIT_RE = re.compile(r"\s*(>>|>)\s*")


def split_beats(normalized: str) -> list[tuple[str, bool]]:
    """Split notation on `>` and `>>`. Return [(beat_text, is_gap_after), ...].

    The is_gap_after flag reflects the separator that FOLLOWS this beat (`>>` = True).
    The final beat has is_gap_after=False.
    """
    if not normalized:
        return []
    parts = _BEAT_SPLIT_RE.split(normalized)
    beats: list[tuple[str, bool]] = []
    i = 0
    while i < len(parts):
        beat_text = parts[i].strip()
        i += 1
        sep = parts[i].strip() if i < len(parts) else ""
        if i < len(parts):
            i += 1
        if beat_text:
            beats.append((beat_text, sep == ">>"))
    return beats


_FLAG_RE = re.compile(r"\(([A-Za-z][^)]*)\)")


def parse_beat(text: str) -> Beat:
    """Parse one beat segment into a structured Beat object."""
    beat = Beat(raw=text)

    # Extract all parens contents; classify lowercase → pre-state, uppercase → component flag.
    flags_found = _FLAG_RE.findall(text)
    stripped = _FLAG_RE.sub("", text).strip()
    stripped = re.sub(r"\s+", " ", stripped)

    for f in flags_found:
        f_clean = f.strip()
        f_upper = f_clean.upper()
        if f_upper in ALL_COMPONENT_FLAGS:
            beat.component_flags.append(f_upper)
        elif f_clean.lower() in PRE_STATE_FLAGS:
            beat.pre_state.append(f_clean.lower())
        elif f_clean.lower() == "op":
            # Some bracketed-style notation uses (op) — operational marker; treat as positional
            beat.component_flags.append("OP")
        else:
            beat.unresolved.append(f"({f_clean})")

    # Now classify the remaining non-paren tokens.
    if not stripped:
        return beat

    tokens = stripped.lower().split()

    # Strip leading "or" connector ("SAME or OP OUT" — early bracket-style) by taking the
    # second branch (Op) as canonical; recorded under unresolved for transparency.
    if "or" in tokens:
        idx = tokens.index("or")
        # Take the post-"or" branch as primary; flag the pre-"or" tokens as alt-side.
        if idx > 0 and idx + 1 < len(tokens):
            alt_tokens = tokens[:idx]
            tokens = tokens[idx + 1:]
            beat.unresolved.append(f"alt-side:{' '.join(alt_tokens)}")

    # Side
    if tokens and tokens[0] in SIDE_PREFIXES:
        beat.side = tokens.pop(0)

    # Direction (sometimes; only when followed by a contact or whirl/swirl)
    if tokens and tokens[0] in DIRECTION_TOKENS:
        beat.direction = tokens.pop(0)

    # Whirl/swirl bigram (e.g. "front whirl") — already consumed direction; check next is "whirl"/"swirl"
    if tokens and tokens[0] in {"whirl", "swirl"}:
        beat.qualifier = f"{beat.direction + ' ' if beat.direction else ''}{tokens.pop(0)}"
        beat.direction = None

    # Body action (spin/duck/dive)
    if tokens and tokens[0] in BODY_ACTIONS:
        beat.base_token = tokens.pop(0)

    # Contact (toe/clip/set/flapper/...)
    if tokens and tokens[0] in CONTACT_TOKENS:
        beat.base_token = tokens.pop(0)

    # Anything left = unresolved
    if tokens:
        beat.unresolved.append(" ".join(tokens))

    return beat


def parse_notation(raw: str) -> ParseResult:
    """Parse a raw notation string into a ParseResult."""
    result = ParseResult()
    if not raw or not raw.strip():
        result.confidence = "none"
        return result
    normalized = normalize_notation(raw)
    beats = split_beats(normalized)
    if not beats:
        result.confidence = "none"
        return result
    for i, (text, gap_after) in enumerate(beats):
        beat = parse_beat(text)
        beat.is_gap_after = gap_after
        result.beats.append(beat)

    result.unresolved_tokens = [u for b in result.beats for u in b.unresolved]

    # Confidence: high if no unresolved tokens; medium if 1; low if 2+; none if no beats parsed.
    if not result.beats:
        result.confidence = "none"
    elif not result.unresolved_tokens:
        result.confidence = "high"
    elif len(result.unresolved_tokens) <= 1:
        result.confidence = "medium"
    else:
        result.confidence = "low"
    return result


# ─────────────────────────────────────────────────────────────────────────
# Derivations (heuristic; diagnostic-only)
# ─────────────────────────────────────────────────────────────────────────

def derive_adds(result: ParseResult) -> int:
    """Additive ADD heuristic.

    Per task brief: DEX=+1, DEL=+1, XBD=+1, BOD=+1, PDX=+1 (when independent), XDEX=+1.
    UNS=0. OP=0. Unresolved=0.

    "Independent" for PDX: only count when PDX is the *primary* flag on a beat, OR when it
    appears on a beat that also has DEX (Paradox-Dex) — i.e. always count it; conservative
    naive reading. The task brief says "only when clearly independent" but offers no
    operational test; the additive-heuristic baseline treats PDX as a flat +1 contributor
    whenever it appears in the component-flag stack. Source-vs-derived delta surfaces any
    over-counting cleanly without forcing perfect match.
    """
    total = 0
    for flag in result.all_component_flags:
        total += COMPONENT_FLAGS_ADD.get(flag, 0)
    return total


def derive_topology_family(result: ParseResult, name: str = "") -> tuple[str, str]:
    """Cluster by symbolic topology. Returns (narrow_family, broad_family).

    Narrow examples (per task brief §5):
    - toe-dex-toe
    - toe-double-dex-toe
    - clip-dex-clip
    - clip-spin-clip
    - clip-whirl-clip
    - butterfly-walk-family
    - torque-family-symbolic
    - osis/dyno spin-family
    - double-over-down family
    - eggbeater family
    - unusual-surface family

    NOT IFPA family. Pure topology clustering on parsed beats.
    """
    if not result.beats:
        return ("no-notation", "no-notation")

    name_lower = (name or "").lower()
    start = result.beats[0].base_token or "?"
    end = result.beats[-1].base_token or "?"
    qualifiers = [b.qualifier for b in result.beats if b.qualifier]
    body_actions = [b.base_token for b in result.beats if b.base_token in BODY_ACTIONS]
    dex_count = sum(1 for b in result.beats if "DEX" in b.component_flags)
    xdex_present = any("XDEX" in b.component_flags for b in result.beats)
    bod_present = any("BOD" in b.component_flags for b in result.beats)
    pdx_present = any("PDX" in b.component_flags for b in result.beats)
    xbd_present = any("XBD" in b.component_flags for b in result.beats)
    uns_present = any("UNS" in b.component_flags for b in result.beats)

    # Unusual-surface override
    if uns_present:
        return (f"unusual-surface-{start}-to-{end}", "unusual-surface-family")

    # Body-action families (spin / duck / dive)
    if body_actions:
        action = body_actions[0]
        if "whirl" in " ".join(qualifiers):
            return (f"{start}-{action}-whirl", f"{action}-family")
        return (f"{start}-{action}-{end}", f"{action}-family")

    # Whirl/swirl families
    if qualifiers:
        q = qualifiers[0]
        if "whirl" in q:
            return (f"{start}-{q.replace(' ', '-')}-{end}", "whirl-family")
        if "swirl" in q:
            return (f"{start}-{q.replace(' ', '-')}-{end}", "swirl-family")

    # Walking-family heuristic: any name containing 'walk' OR start=clip,end=clip with PDX-light dex chain
    if "walk" in name_lower:
        return (f"{start}-walk-{end}", "butterfly-walk-family")

    # Dex chains
    if dex_count == 1 and not bod_present and not pdx_present:
        return (f"{start}-single-dex-{end}", "single-dex-family")
    if dex_count == 2 and not bod_present and not xdex_present:
        return (f"{start}-double-dex-{end}", "double-dex-family")
    if dex_count >= 3:
        return (f"{start}-multi-dex-{end}", "multi-dex-family")
    if xdex_present:
        return (f"{start}-xdex-{end}", "xdex-family")
    if pdx_present:
        return (f"{start}-paradox-{end}", "paradox-family")

    return (f"{start}-to-{end}", "uncategorized")


# Operator-tag derivation (per task brief §6)
def derive_operator_tags(result: ParseResult, name: str, technical: str) -> str:
    """Pipe-delimited operator tags."""
    tags = []
    name_blob = f"{name} {technical}".lower()
    text_all = " ".join(b.raw.lower() for b in result.beats)

    if any(b.base_token == "spin" for b in result.beats) or "spin" in name_blob:
        tags.append("spinning")
    if any(b.base_token == "duck" for b in result.beats) or "duck" in name_blob:
        tags.append("ducking")
    if any(b.base_token == "dive" for b in result.beats) or "dive" in name_blob:
        tags.append("diving")
    if any("PDX" in b.component_flags for b in result.beats) or "paradox" in name_blob:
        tags.append("paradox")
    if any("no plant while" in b.pre_state for b in result.beats) or "symposium" in name_blob:
        tags.append("symposium-no-plant")
    if any(b.side == "same" for b in result.beats):
        tags.append("same-side-segment")
    if any(b.side == "op" for b in result.beats):
        tags.append("opposite-side-segment")
    if any("XBD" in b.component_flags for b in result.beats):
        tags.append("xbd")
    if any("UNS" in b.component_flags for b in result.beats):
        tags.append("unusual-surface")
    if result.beats:
        first = result.beats[0].base_token
        if first == "toe":
            tags.append("toe-start")
        elif first == "clip":
            tags.append("clip-start")
        elif first == "set":
            tags.append("set-start")
        last = result.beats[-1].base_token
        if last == "toe":
            tags.append("toe-finish")
        elif last == "clip":
            tags.append("clip-finish")
        elif last == "flapper":
            tags.append("flapper-finish")
    dex_count = sum(1 for b in result.beats if "DEX" in b.component_flags)
    if dex_count >= 2:
        tags.append("multi-dex")
    if any("XDEX" in b.component_flags for b in result.beats):
        tags.append("xdex")
    # Recursive set-name: notation begins with a non-contact named token (e.g. "Frigidosis", "Dragon")
    if result.beats and result.beats[0].base_token is None and result.beats[0].unresolved:
        tags.append("recursive-set-name")
    return "|".join(tags)


def beat_to_symbol(beat: Beat) -> str:
    """Compact symbolic representation of one beat for parsed_symbol_sequence."""
    parts = []
    if beat.pre_state:
        parts.append(f"[{'|'.join(beat.pre_state)}]")
    if beat.side:
        parts.append(beat.side)
    if beat.direction:
        parts.append(beat.direction)
    if beat.qualifier:
        parts.append(beat.qualifier)
    if beat.base_token:
        parts.append(beat.base_token)
    out = " ".join(parts) if parts else (beat.raw if not beat.component_flags else "")
    if beat.component_flags:
        out += "(" + ",".join(beat.component_flags) + ")"
    return out.strip()


def render_parsed_sequence(result: ParseResult) -> str:
    if not result.beats:
        return ""
    pieces = []
    for i, b in enumerate(result.beats):
        pieces.append(beat_to_symbol(b))
        if i < len(result.beats) - 1:
            pieces.append("⟫" if b.is_gap_after else "›")
    return " ".join(pieces)


# ─────────────────────────────────────────────────────────────────────────
# Beat-role classification (uptime / midtime / downtime)
# ─────────────────────────────────────────────────────────────────────────

def classify_beat_roles(result: ParseResult) -> dict:
    """Sort parsed beats into uptime_dex_1/2, midtime_body, downtime_dexes, final_contact.

    Heuristic:
    - First contact (toe/clip/set): start_contact
    - Last contact (with DEL or terminal): final_contact
    - Body-action beats: midtime_body_actions
    - Dex beats between start and any body-action: uptime
    - Dex beats between body-action and final: downtime
    - If no body-action, dex beats partitioned by ordinal (first half uptime, second half downtime)
    """
    roles = {
        "start_contact": "",
        "uptime_dex_1": "",
        "uptime_dex_2": "",
        "uptime_state_flags": "",
        "midtime_body_actions": "",
        "downtime_dexes": "",
        "final_contact": "",
        "surface_flags": "",
        "positional_flags": "",
    }
    if not result.beats:
        return roles

    # First beat = start contact
    first = result.beats[0]
    if first.base_token in CONTACT_TOKENS:
        roles["start_contact"] = first.base_token
    elif first.base_token is None and first.unresolved:
        # Recursive set-name (e.g. Frigidosis)
        roles["start_contact"] = f"named:{first.unresolved[0]}"
    else:
        roles["start_contact"] = first.raw

    # Last beat = final contact
    last = result.beats[-1]
    if last.base_token in CONTACT_TOKENS:
        roles["final_contact"] = last.base_token
    else:
        roles["final_contact"] = last.raw

    # Identify body-action beats
    body_action_beats = [i for i, b in enumerate(result.beats) if b.base_token in BODY_ACTIONS]
    midtime_bodies = []
    for i in body_action_beats:
        b = result.beats[i]
        label = b.base_token or ""
        if b.pre_state:
            label = f"{'/'.join(b.pre_state)}-{label}"
        midtime_bodies.append(label)
    roles["midtime_body_actions"] = "|".join(midtime_bodies)

    # Dex beats
    dex_beats = [(i, b) for i, b in enumerate(result.beats) if "DEX" in b.component_flags]
    if body_action_beats:
        split = body_action_beats[0]
        uptime_dexes = [b for i, b in dex_beats if i < split]
        downtime_dexes = [b for i, b in dex_beats if i > body_action_beats[-1]]
    else:
        # No body action: split dexes evenly (first half uptime, second half downtime)
        mid = (len(dex_beats) + 1) // 2
        uptime_dexes = [b for i, (idx, b) in enumerate(dex_beats) if i < mid]
        downtime_dexes = [b for i, (idx, b) in enumerate(dex_beats) if i >= mid]

    def fmt_dex(b: Beat) -> str:
        side = b.side or ""
        direction = b.direction or ""
        qual = b.qualifier or ""
        flags = ",".join(f for f in b.component_flags if f not in {"DEX", "DEL"})
        pieces = [p for p in [side, direction, qual] if p]
        out = " ".join(pieces) if pieces else "dex"
        if flags:
            out += f" [{flags}]"
        return out.strip()

    if uptime_dexes:
        roles["uptime_dex_1"] = fmt_dex(uptime_dexes[0])
        if len(uptime_dexes) > 1:
            roles["uptime_dex_2"] = fmt_dex(uptime_dexes[1])
    roles["downtime_dexes"] = "|".join(fmt_dex(b) for b in downtime_dexes)

    # State flags (uptime pre-state collected across uptime beats)
    upstate = set()
    upper_bound = body_action_beats[0] if body_action_beats else len(result.beats) - 1
    for b in result.beats[:upper_bound]:
        for ps in b.pre_state:
            upstate.add(ps)
    roles["uptime_state_flags"] = "|".join(sorted(upstate))

    # Surface flags
    surface = []
    if any("UNS" in b.component_flags for b in result.beats):
        surface.append("UNS")
    roles["surface_flags"] = "|".join(surface)

    # Positional flags
    positional = []
    if any("XBD" in b.component_flags for b in result.beats):
        positional.append("XBD")
    roles["positional_flags"] = "|".join(positional)

    return roles


# ─────────────────────────────────────────────────────────────────────────
# Source readers
# ─────────────────────────────────────────────────────────────────────────

def read_fm_inventory() -> list[dict]:
    out: list[dict] = []
    with FM_INVENTORY.open(newline="") as f:
        for row in csv.DictReader(f):
            out.append({
                "source": "footbagmoves",
                "source_file": row.get("source_paste_file") or "",
                "move_name": row.get("display_name") or "",
                "alternate_names": row.get("aliases_raw") or "",
                "technical_name": row.get("technical_name") or "",
                "source_adds": row.get("add_count") or "",
                "symbolic_notation_raw": row.get("operational_notation_raw") or "",
                "_external_id": row.get("synthetic_external_id") or "",
                "_same_side_variant": row.get("same_side_variant") or "0",
            })
    return out


def read_passback_queue() -> list[dict]:
    out: list[dict] = []
    if not PASSBACK_QUEUE.exists():
        return out
    with PASSBACK_QUEUE.open(newline="") as f:
        for row in csv.DictReader(f):
            out.append({
                "source": "passback",
                "source_file": "passback_review_queue.csv",
                "move_name": row.get("raw_name") or "",
                "alternate_names": "",
                "technical_name": row.get("sort_friendly") or "",
                "source_adds": row.get("adds") or "",
                "symbolic_notation_raw": "",  # PassBack queue doesn't carry set-arc notation
                "_player_name": row.get("player_name") or "",
                "_date": row.get("date_recorded") or "",
                "_url": row.get("url") or "",
            })
    return out


# ─────────────────────────────────────────────────────────────────────────
# Row builder
# ─────────────────────────────────────────────────────────────────────────

def build_master_row(src: dict) -> dict:
    raw_notation = src.get("symbolic_notation_raw") or ""
    result = parse_notation(raw_notation)
    roles = classify_beat_roles(result)
    derived = derive_adds(result)
    component_flag_stack = "|".join(result.all_component_flags)
    narrow_family, broad_family = derive_topology_family(result, name=src.get("move_name", ""))
    op_tags = derive_operator_tags(result, src.get("move_name", ""), src.get("technical_name", ""))
    parsed_seq = render_parsed_sequence(result)
    unresolved = "|".join(result.unresolved_tokens)

    # Delta
    try:
        src_adds = int(src.get("source_adds") or 0)
    except ValueError:
        src_adds = 0
    delta = src_adds - derived if raw_notation else None

    # Review status
    if not raw_notation.strip():
        review_status = "source_has_no_notation"
    elif result.confidence == "high" and delta in (None, 0):
        review_status = "ok"
    elif result.unresolved_tokens or result.confidence == "low":
        review_status = "needs_review"
    else:
        review_status = "ok"

    # Notes (lightweight provenance)
    notes_parts = []
    if src.get("_same_side_variant") == "1":
        notes_parts.append("fm:same-side-variant")
    if src.get("_player_name"):
        notes_parts.append(f"passback:{src['_player_name']}/{src.get('_date','')}")
    if delta not in (None, 0) and raw_notation:
        notes_parts.append(f"add-delta:{delta:+d}")

    return {
        "source": src["source"],
        "source_file": src["source_file"],
        "move_name": src["move_name"],
        "alternate_names": src["alternate_names"],
        "technical_name": src["technical_name"],
        "source_adds": src["source_adds"],
        "symbolic_notation_raw": raw_notation,
        "start_contact": roles["start_contact"],
        "uptime_dex_1": roles["uptime_dex_1"],
        "uptime_dex_2": roles["uptime_dex_2"],
        "uptime_state_flags": roles["uptime_state_flags"],
        "midtime_body_actions": roles["midtime_body_actions"],
        "downtime_dexes": roles["downtime_dexes"],
        "final_contact": roles["final_contact"],
        "surface_flags": roles["surface_flags"],
        "positional_flags": roles["positional_flags"],
        "component_flags": component_flag_stack,
        "parsed_symbol_sequence": parsed_seq,
        "derived_adds": derived if raw_notation else "",
        "source_vs_derived_delta": "" if delta is None else delta,
        "derived_symbolic_family": narrow_family,
        "derived_topology_family": broad_family,
        "unresolved_tokens": unresolved,
        "parse_confidence": result.confidence,
        "review_status": review_status,
        "notes": "|".join(notes_parts),
        "_op_tags": op_tags,  # tucked-in op tags for analysis; not in main FIELDS but included for downstream summary
    }


# ─────────────────────────────────────────────────────────────────────────
# Output writers
# ─────────────────────────────────────────────────────────────────────────

def write_csv(rows: list[dict], path: Path, fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fieldnames})


def write_xlsx(rows: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Symbolic Grammar Master"
    ws.append(FIELDS)
    # Header style
    header_fill = PatternFill(start_color="FFE7E0EC", end_color="FFE7E0EC", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in rows:
        ws.append([row.get(k, "") for k in FIELDS])
    # Freeze header row
    ws.freeze_panes = "A2"
    # Column widths (sane defaults)
    widths = {
        "source": 14, "source_file": 24, "move_name": 28, "alternate_names": 24,
        "technical_name": 36, "source_adds": 11, "symbolic_notation_raw": 64,
        "start_contact": 14, "uptime_dex_1": 18, "uptime_dex_2": 18,
        "uptime_state_flags": 18, "midtime_body_actions": 22, "downtime_dexes": 28,
        "final_contact": 14, "surface_flags": 14, "positional_flags": 16,
        "component_flags": 24, "parsed_symbol_sequence": 64, "derived_adds": 13,
        "source_vs_derived_delta": 22, "derived_symbolic_family": 28,
        "derived_topology_family": 28, "unresolved_tokens": 32, "parse_confidence": 16,
        "review_status": 22, "notes": 32,
    }
    for col_idx, name in enumerate(FIELDS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = widths.get(name, 16)
    wb.save(path)


def write_unresolved_tokens_csv(rows: list[dict]) -> None:
    counter: Counter[str] = Counter()
    example_rows: dict[str, str] = {}
    for row in rows:
        if not row["unresolved_tokens"]:
            continue
        for tok in row["unresolved_tokens"].split("|"):
            if not tok.strip():
                continue
            counter[tok] += 1
            if tok not in example_rows:
                example_rows[tok] = row["move_name"]
    UNRESOLVED_CSV.parent.mkdir(parents=True, exist_ok=True)
    with UNRESOLVED_CSV.open("w", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(["unresolved_token", "occurrence_count", "example_move_name"])
        for tok, n in counter.most_common():
            w.writerow([tok, n, example_rows.get(tok, "")])


def write_family_summary_csv(rows: list[dict]) -> None:
    narrow: Counter[str] = Counter()
    broad: Counter[str] = Counter()
    family_examples: dict[str, list[str]] = {}
    for row in rows:
        nf = row["derived_symbolic_family"]
        bf = row["derived_topology_family"]
        narrow[nf] += 1
        broad[bf] += 1
        family_examples.setdefault(bf, []).append(row["move_name"])
    FAMILY_CSV.parent.mkdir(parents=True, exist_ok=True)
    with FAMILY_CSV.open("w", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_ALL)
        w.writerow(["scope", "family_slug", "row_count", "example_moves"])
        for fam, n in broad.most_common():
            examples = ", ".join(family_examples[fam][:5])
            w.writerow(["topology_family", fam, n, examples])
        w.writerow([])  # blank separator
        for fam, n in narrow.most_common():
            w.writerow(["symbolic_family", fam, n, ""])


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────

def main() -> int:
    fm_rows = read_fm_inventory()
    pb_rows = read_passback_queue()
    all_sources = fm_rows + pb_rows
    rows = [build_master_row(src) for src in all_sources]

    write_csv(rows, MASTER_CSV, FIELDS)
    write_xlsx(rows, MASTER_XLSX)
    write_unresolved_tokens_csv(rows)
    write_family_summary_csv(rows)

    # Summary to stdout
    n_total = len(rows)
    n_with_notation = sum(1 for r in rows if r["symbolic_notation_raw"])
    parse_counts = Counter(r["parse_confidence"] for r in rows)
    review_counts = Counter(r["review_status"] for r in rows)
    print(f"Total rows         : {n_total}")
    print(f"  with notation    : {n_with_notation}")
    print(f"  without notation : {n_total - n_with_notation}")
    print(f"Parse confidence   : {dict(parse_counts)}")
    print(f"Review status      : {dict(review_counts)}")
    print(f"FM rows            : {len(fm_rows)}")
    print(f"PassBack rows      : {len(pb_rows)}")
    print(f"Wrote: {MASTER_CSV}")
    print(f"Wrote: {MASTER_XLSX}")
    print(f"Wrote: {UNRESOLVED_CSV}")
    print(f"Wrote: {FAMILY_CSV}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
